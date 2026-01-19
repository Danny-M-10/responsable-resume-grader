"""
Streamlit Web Interface for Candidate Ranker Application

Run with: streamlit run app.py
"""

import streamlit as st
import streamlit.components.v1 as components
import os
import tempfile
import csv
import io
import base64
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

# Configure logging
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
from candidate_ranker import CandidateRankerApp
from config import OpenAIConfig
from ai_job_parser import AIJobParser
from loading_components import get_cycling_message
from db import init_db, get_db, utcnow_str
from auth import (
    create_user, authenticate, get_user_by_session, destroy_session, update_session_activity,
    request_password_reset, reset_password_with_code, reset_password_with_token
)
from storage import save_bytes
from models import CandidateScore
from utils import prepare_query, is_safe_path, MAX_PDF_SIZE
from resume_database import (
    save_candidate_profile, get_candidate_profile, search_candidates,
    filter_candidates, update_candidate_profile, delete_candidate_profile,
    get_candidates_for_analysis, link_candidate_to_analysis, get_candidate_analyses
)
# Theme system imports (needed early for _auth_gate)
from ui.components import get_initial_theme, inject_theme_css, render_theme_toggle, render_theme_script, set_theme


@st.cache_resource
def _bootstrap_db():
    init_db()
    return True


def set_session_token_cookie(token):
    """Set session token in browser cookie using JavaScript"""
    if token:
        cookie_js = f"""
        <script>
        const expires = new Date();
        expires.setTime(expires.getTime() + (7 * 24 * 60 * 60 * 1000)); // 7 days
        document.cookie = "session_token={token}; path=/; expires=" + expires.toUTCString() + "; SameSite=Lax";
        </script>
        """
    else:
        # Clear cookie
        cookie_js = """
        <script>
        document.cookie = "session_token=; path=/; expires=Thu, 01 Jan 1970 00:00:00 UTC;";
        </script>
        """
    # Note: components.html doesn't support key parameter in Streamlit
    components.html(cookie_js, height=0, width=0)


def _auth_gate():
    """
    Simple login/register gate. Returns (user_id, email) when authenticated.
    Renders forms and stops execution if not authenticated.
    """
    _bootstrap_db()

    # Initialize session token from session state or query params
    # Streamlit query params can persist across reloads
    token = st.session_state.get("session_token")
    
    # If no token in session state, try to get from query params (fallback)
    if not token:
        query_params = st.query_params
        # Handle both dict-like and list-like query param formats
        if hasattr(query_params, 'get'):
            token_param = query_params.get("token")
            if token_param:
                # If it's a list, get first element; otherwise use directly
                token = token_param[0] if isinstance(token_param, list) else token_param
                if token:
                    st.session_state["session_token"] = token
    
    # Also set cookie for better persistence (runs on every request)
    if token:
        set_session_token_cookie(token)
    
    user = get_user_by_session(token) if token else None

    def logout():
        if st.session_state.get("session_token"):
            destroy_session(st.session_state["session_token"])
        st.session_state["session_token"] = None
        st.session_state.pop("user_id", None)
        st.session_state.pop("user_email", None)
        # Clear cookie
        set_session_token_cookie(None)
        # Clear query param
        st.query_params.clear()
        st.success("Logged out.")
        st.rerun()

    if user:
        user_id, email = user
        st.session_state["user_id"] = user_id
        st.session_state["user_email"] = email
        # Ensure token is stored in cookie for persistence
        if token:
            set_session_token_cookie(token)
            update_session_activity(token)
        # Theme toggle in sidebar header
        # Get current theme (will be initialized in main() but we need it here)
        current_theme = get_initial_theme()
        st.sidebar.markdown("### Theme")
        render_theme_toggle(current_theme)
        st.sidebar.markdown("---")
        
        st.sidebar.success(f"Logged in as {email}")
        st.sidebar.info("Session expires after 2 hours of inactivity")
        
        # Previous Analyses section in sidebar
        st.sidebar.markdown("---")
        st.sidebar.markdown("### Previous Analyses")
        analyses = load_user_analyses(user_id)
        if analyses:
            st.sidebar.info(f"You have {len(analyses)} previous analysis(es)")
            # Show most recent 5 in sidebar
            for analysis in analyses[:5]:
                with st.sidebar.expander(f"{analysis['title'][:30]}...", expanded=False):
                    st.caption(f"Date: {analysis['created_at'][:10]}")
                    st.caption(f"Location: {analysis['location']}")
                    st.caption(f"Candidates: {analysis['num_candidates']}")
                    if st.button("View", key=f"sidebar_view_{analysis['report_id']}"):
                        st.session_state["view_report_id"] = analysis['report_id']
                        st.rerun()
        else:
            st.sidebar.info("No previous analyses yet")
        
        if st.sidebar.button("Logout"):
            logout()
        return user

    st.sidebar.info("Login required to continue.")
    
    # Check if we should show login tab (after registration)
    default_tab = st.session_state.get("active_tab", "Login")
    tab_login, tab_register = st.tabs(["Login", "Register"])
    
    # Pre-fill email if coming from registration
    registration_email = st.session_state.get("registration_email", "")

    with tab_login:
        login_email = st.text_input("Email", value=registration_email, key="login_email")
        login_password = st.text_input("Password", type="password", key="login_password")
        
        # Forgot Password link
        col1, col2 = st.columns([3, 1])
        with col2:
            if st.button("Forgot Password?", key="forgot_password_link"):
                st.session_state["show_password_reset"] = True
        
        if st.session_state.get("show_password_reset", False):
            st.markdown("---")
            st.markdown("### Password Reset")
            
            reset_method = st.radio(
                "Reset method:",
                ["Send reset code (6-digit)", "Use reset token"],
                key="reset_method"
            )
            
            reset_email = st.text_input("Email", key="reset_email")
            
            if reset_method == "Send reset code (6-digit)":
                if st.button("Send Reset Code", key="send_code"):
                    try:
                        code, user_id = request_password_reset(reset_email)
                        st.session_state["reset_code"] = code
                        st.session_state["reset_email"] = reset_email
                        
                        # Only show code in development mode
                        if os.getenv('ENVIRONMENT', '').lower() == 'development':
                            st.success(f"Reset code generated: **{code}** (Development mode)")
                        else:
                            st.success("Reset code has been sent to your email. Please check your inbox.")
                        st.info("Code expires in 1 hour. Use it below to reset your password.")
                    except Exception as e:
                        st.error(str(e))
                
                if st.session_state.get("reset_code"):
                    st.markdown("---")
                    st.markdown("### Enter Reset Code")
                    entered_code = st.text_input("6-digit code", key="entered_code", max_chars=6)
                    new_password = st.text_input("New Password", type="password", key="new_password_reset")
                    confirm_password = st.text_input("Confirm New Password", type="password", key="confirm_password_reset")
                    
                    if st.button("Reset Password", key="reset_with_code"):
                        if new_password != confirm_password:
                            st.error("Passwords do not match.")
                        elif not entered_code or not entered_code.isdigit() or len(entered_code) != 6:
                            st.error("Code must be exactly 6 digits.")
                        else:
                            if reset_password_with_code(st.session_state["reset_email"], entered_code, new_password):
                                st.success("Password reset successfully! Please log in.")
                                st.session_state.pop("show_password_reset", None)
                                st.session_state.pop("reset_code", None)
                                st.session_state.pop("reset_email", None)
                                st.rerun()
                            else:
                                st.error("Invalid or expired code. Please request a new code.")
            
            else:  # Use reset token
                reset_token = st.text_input("Reset Token", key="reset_token")
                new_password = st.text_input("New Password", type="password", key="new_password_token")
                confirm_password = st.text_input("Confirm New Password", type="password", key="confirm_password_token")
                
                if st.button("Reset Password", key="reset_with_token"):
                    if new_password != confirm_password:
                        st.error("Passwords do not match.")
                    elif not reset_token:
                        st.error("Please enter the reset token.")
                    else:
                        if reset_password_with_token(reset_token, new_password):
                            st.success("Password reset successfully! Please log in.")
                            st.session_state.pop("show_password_reset", None)
                            st.rerun()
                        else:
                            st.error("Invalid or expired token.")
            
            if st.button("Cancel", key="cancel_reset"):
                st.session_state.pop("show_password_reset", None)
                st.session_state.pop("reset_code", None)
                st.session_state.pop("reset_email", None)
                st.rerun()
        
        if st.button("Login", key="login_button"):
            try:
                session_token, user_id = authenticate(login_email, login_password)
                st.session_state["session_token"] = session_token
                st.session_state["user_id"] = user_id
                st.session_state["user_email"] = login_email.strip().lower()
                # Store token in cookie and query params for persistence
                set_session_token_cookie(session_token)
                st.query_params["token"] = session_token
                # Clear registration email after successful login
                st.session_state.pop("registration_email", None)
                st.session_state.pop("active_tab", None)
                st.success("Logged in successfully.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with tab_register:
        reg_email = st.text_input("Email", key="reg_email")
        reg_password = st.text_input("Password", type="password", key="reg_password")
        reg_confirm = st.text_input("Confirm Password", type="password", key="reg_confirm")
        if st.button("Register", key="register_button"):
            if reg_password != reg_confirm:
                st.error("Passwords do not match.")
            elif not reg_email or not reg_password:
                st.error("Email and password are required.")
            else:
                try:
                    user_id = create_user(reg_email, reg_password)
                    # Store email for pre-filling login
                    st.session_state["registration_email"] = reg_email.strip().lower()
                    st.session_state["active_tab"] = "Login"
                    st.success("Account created successfully! Please log in.")
                    st.balloons()
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    st.stop()


# Use shared utility function instead of local wrapper
_prepare_query_wrapper = prepare_query


def load_user_analyses(user_id: str) -> List[Dict]:
    """
    Load list of user's previous analyses from database.
    Returns list of dicts with report info.
    """
    if not user_id:
        logger.error("load_user_analyses called with empty user_id")
        return []
    
    try:
        with get_db() as conn:
            cur = conn.cursor()
            query = _prepare_query_wrapper(conn, """
                SELECT r.id, r.created_at, r.pdf_path, r.summary_json,
                       j.title, j.location, j.certifications_json
                FROM reports r
                JOIN job_descriptions j ON r.job_description_id = j.id
                WHERE r.user_id = ?
                ORDER BY r.created_at DESC
            """)
            cur.execute(query, (user_id,))
            rows = cur.fetchall()
            
            analyses = []
            for row in rows:
                report_id, created_at, pdf_path, summary_json, title, location, certs_json = row
                try:
                    summary = json.loads(summary_json) if summary_json else {}
                except (json.JSONDecodeError, TypeError) as e:
                    logger.warning(f"Error parsing summary_json for report {report_id}: {e}")
                    summary = {}
                analyses.append({
                    'report_id': report_id,
                    'created_at': created_at,
                    'pdf_path': pdf_path,
                    'title': title or 'Untitled',
                    'location': location or 'Not specified',
                    'num_candidates': summary.get('all_candidates', 0),
                    'top_candidates': summary.get('top_candidates', 0),
                })
            logger.info(f"Successfully loaded {len(analyses)} analyses for user {user_id}")
            return analyses
    except Exception as e:
        logger.error(f"Error loading user analyses for user {user_id}: {e}", exc_info=True)
        return []


def load_analysis_data(report_id: str, user_id: str) -> Optional[Dict]:
    """
    Load full analysis data for a specific report.
    Returns dict with job_details, candidate_scores, pdf_data, etc.
    """
    if not report_id or not user_id:
        logger.error(f"load_analysis_data called with invalid params: report_id={report_id}, user_id={user_id}")
        return None
    
    try:
        with get_db() as conn:
            cur = conn.cursor()
            
            # Get report and job description
            query = _prepare_query_wrapper(conn, """
                SELECT r.id, r.created_at, r.pdf_path, r.summary_json,
                       j.id as job_id, j.title, j.location, j.certifications_json,
                       j.required_skills_json, j.preferred_skills_json, j.full_description
                FROM reports r
                JOIN job_descriptions j ON r.job_description_id = j.id
                WHERE r.id = ? AND r.user_id = ?
            """)
            cur.execute(query, (report_id, user_id))
            row = cur.fetchone()
            
            if not row:
                return None
            
            (report_id_db, created_at, pdf_path, summary_json, job_id, title, location,
             certs_json, req_skills_json, pref_skills_json, full_desc) = row
            
            # Get candidate scores
            query = _prepare_query_wrapper(conn, """
                SELECT candidate_name, email, phone, fit_score, rationale, raw_json
                FROM candidate_scores
                WHERE report_id = ?
                ORDER BY fit_score DESC
            """)
            cur.execute(query, (report_id,))
            score_rows = cur.fetchall()
            
            # Reconstruct CandidateScore objects
            candidate_scores = []
            for score_row in score_rows:
                name, email, phone, fit_score, rationale, raw_json = score_row
                if raw_json:
                    try:
                        cand_data = json.loads(raw_json)
                        # Reconstruct CandidateScore from stored data
                        candidate_scores.append(CandidateScore(
                            name=cand_data.get('name', name),
                            phone=cand_data.get('phone', phone or ''),
                            email=cand_data.get('email', email or ''),
                            certifications=cand_data.get('certifications', []),
                            fit_score=float(fit_score),
                            chain_of_thought=cand_data.get('chain_of_thought', ''),
                            rationale=rationale or cand_data.get('rationale', ''),
                            experience_match=cand_data.get('experience_match', {}),
                            certification_match=cand_data.get('certification_match', {}),
                            skills_match=cand_data.get('skills_match', {}),
                            location_match=cand_data.get('location_match', False),
                            component_scores=cand_data.get('component_scores', {}),
                            calibration_applied=cand_data.get('calibration_applied', False),
                            calibration_factor=cand_data.get('calibration_factor', 1.0),
                        ))
                    except (json.JSONDecodeError, KeyError, ValueError):
                        # Fallback to basic data if JSON parsing fails
                        candidate_scores.append(CandidateScore(
                            name=name or 'Unknown',
                            phone=phone or '',
                            email=email or '',
                            certifications=[],
                            fit_score=float(fit_score),
                            chain_of_thought='',
                            rationale=rationale or '',
                            experience_match={},
                            certification_match={},
                            skills_match={},
                            location_match=False,
                        ))
            
            # Load PDF data if file exists (with size and path validation)
            # Note: PDF loading failures should not prevent the analysis from being displayed
            pdf_data = None
            if pdf_path:
                if os.path.exists(pdf_path):
                    try:
                        # Validate path safety
                        if is_safe_path(pdf_path):
                            # Check file size
                            file_size = os.path.getsize(pdf_path)
                            if file_size <= MAX_PDF_SIZE:
                                with open(pdf_path, 'rb') as f:
                                    pdf_data = f.read()
                            else:
                                logger.warning(f"PDF file too large: {file_size} bytes (max: {MAX_PDF_SIZE})")
                        else:
                            logger.warning(f"Unsafe PDF path detected: {pdf_path}")
                    except Exception as e:
                        logger.warning(f"Failed to load PDF from {pdf_path}: {e}", exc_info=True)
                        # Continue without PDF data - don't fail the entire load
                else:
                    # Only log warning if PDF is recent (within last 6 hours) to avoid spam from old deleted PDFs
                    # Use debug level for missing PDFs to reduce log noise - missing PDFs are handled gracefully
                    try:
                        # Try to extract timestamp from filename
                        import re
                        match = re.search(r'(\d{8}_\d{6})', pdf_path)
                        if match:
                            pdf_timestamp_str = match.group(1)
                            pdf_date = datetime.strptime(pdf_timestamp_str, "%Y%m%d_%H%M%S")
                            hours_old = (datetime.now() - pdf_date).total_seconds() / 3600
                            # Only log as warning if very recent (within 6 hours), otherwise debug
                            if hours_old < 6:
                                logger.warning(f"PDF file not found at path: {pdf_path} (created {hours_old:.1f} hours ago)")
                            else:
                                logger.debug(f"PDF file not found at path: {pdf_path} (created {hours_old:.1f} hours ago - likely cleaned up)")
                        else:
                            # If we can't parse timestamp, log as debug (likely an old file or non-standard path)
                            logger.debug(f"PDF file not found at path: {pdf_path} (unable to parse timestamp)")
                    except Exception:
                        # If parsing fails, don't log - likely an old file
                        pass
                    # Continue without PDF data - file might have been moved/deleted
            
            # Parse job details
            certifications = []
            if certs_json:
                try:
                    certs_data = json.loads(certs_json)
                    certifications = [c for c in certs_data if isinstance(c, dict)]
                except json.JSONDecodeError:
                    pass
            
            required_skills = []
            if req_skills_json:
                try:
                    required_skills = json.loads(req_skills_json)
                except json.JSONDecodeError:
                    pass
            
            preferred_skills = []
            if pref_skills_json:
                try:
                    preferred_skills = json.loads(pref_skills_json)
                except json.JSONDecodeError:
                    pass
            
            return {
                'report_id': report_id_db,
                'created_at': created_at,
                'pdf_path': pdf_path,
                'pdf_data': pdf_data,
                'candidate_scores': candidate_scores,
                'job_details': {
                    'title': title or 'Untitled',
                    'location': location or 'Not specified',
                    'certifications': certifications,
                    'full_description': full_desc or '',
                },
                'processing_time': 0,  # Not stored, use 0 as default
                'timestamp': created_at,
            }
    except Exception as e:
        logger.error(f"Error loading analysis data for report {report_id}, user {user_id}: {e}", exc_info=True)
        return None


def init_session_state():
    """Initialize session state variables for result persistence"""
    if 'results' not in st.session_state:
        st.session_state.results = None
    if 'processing' not in st.session_state:
        st.session_state.processing = False
    if 'show_results_page' not in st.session_state:
        st.session_state.show_results_page = False
    if 'last_job_title' not in st.session_state:
        st.session_state.last_job_title = ""
    if 'show_history_list' not in st.session_state:
        st.session_state.show_history_list = False


def clear_results():
    """Clear stored results to start fresh"""
    st.session_state.results = None
    st.session_state.processing = False
    st.session_state.show_results_page = False


def generate_csv_export(candidate_scores, job_title, location):
    """Generate enhanced CSV data from candidate scores with all details"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Job details header
    writer.writerow(['Job Title', job_title])
    writer.writerow(['Location', location])
    writer.writerow([])  # Empty row
    
    # Enhanced header row with component scores
    writer.writerow([
        'Rank', 'Name', 'Email', 'Phone', 'Fit Score',
        'Must-Have Certs Match', 'Bonus Certs Match', 'Required Skills Match Rate',
        'Preferred Skills Match Rate', 'Experience Level Match', 'Job Title Match',
        'Location Match', 'Certifications', 'Skills', 'Years Experience',
        'Rationale'
    ])

    # Sort candidates by score
    sorted_candidates = sorted(candidate_scores, key=lambda x: x.fit_score, reverse=True)

    # Data rows with enhanced details
    for i, candidate in enumerate(sorted_candidates, 1):
        cert_match = candidate.certification_match
        skills_match = candidate.skills_match
        exp_match = candidate.experience_match
        
        writer.writerow([
            i,
            candidate.name,
            candidate.email,
            candidate.phone,
            f"{candidate.fit_score:.2f}",
            'Yes' if cert_match.get('has_must_have', False) else 'No',
            'Yes' if cert_match.get('has_bonus', False) else 'No',
            f"{skills_match.get('required_match_rate', 0.0):.2%}" if skills_match else 'N/A',
            f"{skills_match.get('preferred_match_rate', 0.0):.2%}" if skills_match else 'N/A',
            f"{exp_match.get('level_match', 0.0):.2%}" if exp_match else 'N/A',
            'Yes' if candidate.location_match else 'No',
            'Yes' if candidate.location_match else 'No',
            '; '.join(candidate.certifications) if candidate.certifications else 'None',
            '; '.join(skills_match.get('candidate_skills', [])) if skills_match else 'None',
            exp_match.get('years', 0) if exp_match else 0,
            candidate.rationale[:1000] if candidate.rationale else ''  # Longer rationale
        ])

    return output.getvalue()


def _matches_date_filter(created_at_str: str, date_from, date_to) -> bool:
    """Check if analysis date matches date filter range"""
    try:
        if created_at_str.endswith('Z'):
            dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
        else:
            dt = datetime.fromisoformat(created_at_str)
        analysis_date = dt.date()
        
        if date_from and analysis_date < date_from:
            return False
        if date_to and analysis_date > date_to:
            return False
        return True
    except (ValueError, AttributeError, TypeError):
        return True  # Include if date parsing fails


def _sort_analyses(analyses: List[Dict], sort_option: str) -> List[Dict]:
    """Sort analyses based on selected option"""
    if sort_option == "Date (Newest First)":
        return sorted(analyses, key=lambda x: x.get('created_at', ''), reverse=True)
    elif sort_option == "Date (Oldest First)":
        return sorted(analyses, key=lambda x: x.get('created_at', ''), reverse=False)
    elif sort_option == "Job Title (A-Z)":
        return sorted(analyses, key=lambda x: x.get('title', '').lower())
    elif sort_option == "Job Title (Z-A)":
        return sorted(analyses, key=lambda x: x.get('title', '').lower(), reverse=True)
    elif sort_option == "Candidates (Most First)":
        return sorted(analyses, key=lambda x: x.get('num_candidates', 0), reverse=True)
    elif sort_option == "Candidates (Least First)":
        return sorted(analyses, key=lambda x: x.get('num_candidates', 0), reverse=False)
    else:
        return analyses


def delete_analysis(report_id: str, user_id: str) -> bool:
    """
    Delete an analysis and its associated data.
    Returns True if successful, False otherwise.
    """
    with get_db() as conn:
        try:
            cur = conn.cursor()
            
            # Verify ownership
            cur.execute(
                _prepare_query_wrapper(conn, "SELECT pdf_path FROM reports WHERE id = ? AND user_id = ?"),
                (report_id, user_id)
            )
            row = cur.fetchone()
            if not row:
                return False  # Not found or not owned by user
            
            pdf_path = row[0]
            
            # Delete from database (cascade will handle related records)
            cur.execute(
                _prepare_query_wrapper(conn, "DELETE FROM reports WHERE id = ? AND user_id = ?"),
                (report_id, user_id)
            )
            
            # Optionally delete PDF file
            if pdf_path and os.path.exists(pdf_path) and is_safe_path(pdf_path):
                try:
                    os.remove(pdf_path)
                except Exception as e:
                    logger.warning(f"Failed to delete PDF file {pdf_path}: {e}", exc_info=True)
                    # Continue even if file deletion fails
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error deleting analysis {report_id}: {e}", exc_info=True)
            conn.rollback()
            return False


def generate_excel_export(analysis_data: Dict, job_title: str, location: str) -> Optional[bytes]:
    """Generate Excel export with multiple sheets"""
    try:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Try pandas as fallback
            try:
                import pandas as pd
                use_pandas = True
            except ImportError:
                logger.warning("Neither openpyxl nor pandas available for Excel export")
                return None
        else:
            use_pandas = False
        
        if use_pandas:
            # Use pandas for Excel export
            import pandas as pd
            from io import BytesIO
            
            # Create Excel writer
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Job Summary
                job_summary = pd.DataFrame({
                    'Field': ['Job Title', 'Location', 'Total Candidates', 'Analysis Date'],
                    'Value': [
                        job_title,
                        location,
                        len(analysis_data.get('candidate_scores', [])),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                job_summary.to_excel(writer, sheet_name='Job Summary', index=False)
                
                # Sheet 2: Candidate Rankings
                candidates_data = []
                for i, candidate in enumerate(sorted(analysis_data.get('candidate_scores', []), 
                                                     key=lambda x: x.fit_score, reverse=True), 1):
                    candidates_data.append({
                        'Rank': i,
                        'Name': candidate.name,
                        'Email': candidate.email,
                        'Phone': candidate.phone,
                        'Fit Score': candidate.fit_score,
                        'Must-Have Certs': 'Yes' if candidate.certification_match.get('has_must_have') else 'No',
                        'Bonus Certs': 'Yes' if candidate.certification_match.get('has_bonus') else 'No',
                        'Required Skills Match': f"{candidate.skills_match.get('required_match_rate', 0.0):.2%}" if candidate.skills_match else 'N/A',
                        'Preferred Skills Match': f"{candidate.skills_match.get('preferred_match_rate', 0.0):.2%}" if candidate.skills_match else 'N/A',
                        'Experience Match': f"{candidate.experience_match.get('level_match', 0.0):.2%}" if candidate.experience_match else 'N/A',
                        'Location Match': 'Yes' if candidate.location_match else 'No',
                        'Certifications': '; '.join(candidate.certifications) if candidate.certifications else 'None',
                        'Years Experience': candidate.experience_match.get('years', 0) if candidate.experience_match else 0,
                        'Rationale': candidate.rationale[:500] if candidate.rationale else ''
                    })
                
                if candidates_data:
                    candidates_df = pd.DataFrame(candidates_data)
                    candidates_df.to_excel(writer, sheet_name='Candidates', index=False)
                
                # Sheet 3: Component Scores Matrix
                if analysis_data.get('candidate_scores'):
                    scores_matrix = []
                    for candidate in sorted(analysis_data['candidate_scores'], 
                                          key=lambda x: x.fit_score, reverse=True):
                        comp_scores = candidate.component_scores or {}
                        row = {'Candidate': candidate.name, 'Fit Score': candidate.fit_score}
                        row.update({
                            f"{k.replace('_', ' ').title()}": v 
                            for k, v in comp_scores.items()
                        })
                        scores_matrix.append(row)
                    
                    if scores_matrix:
                        scores_df = pd.DataFrame(scores_matrix)
                        scores_df.to_excel(writer, sheet_name='Component Scores', index=False)
            
            output.seek(0)
            return output.read()
        else:
            # Use openpyxl directly for better formatting
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Sheet 1: Job Summary
            ws1 = wb.create_sheet("Job Summary")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            ws1['A1'] = 'Field'
            ws1['B1'] = 'Value'
            ws1['A1'].fill = header_fill
            ws1['A1'].font = header_font
            ws1['B1'].fill = header_fill
            ws1['B1'].font = header_font
            
            ws1['A2'] = 'Job Title'
            ws1['B2'] = job_title
            ws1['A3'] = 'Location'
            ws1['B3'] = location
            ws1['A4'] = 'Total Candidates'
            ws1['B4'] = len(analysis_data.get('candidate_scores', []))
            ws1['A5'] = 'Analysis Date'
            ws1['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Sheet 2: Candidates
            ws2 = wb.create_sheet("Candidates")
            headers = ['Rank', 'Name', 'Email', 'Phone', 'Fit Score', 'Must-Have Certs', 
                      'Bonus Certs', 'Req Skills %', 'Pref Skills %', 'Exp Match %', 
                      'Location Match', 'Certifications', 'Years Exp', 'Rationale']
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            sorted_candidates = sorted(analysis_data.get('candidate_scores', []), 
                                     key=lambda x: x.fit_score, reverse=True)
            for row_idx, candidate in enumerate(sorted_candidates, 2):
                ws2.cell(row=row_idx, column=1, value=row_idx - 1)
                ws2.cell(row=row_idx, column=2, value=candidate.name)
                ws2.cell(row=row_idx, column=3, value=candidate.email)
                ws2.cell(row=row_idx, column=4, value=candidate.phone)
                ws2.cell(row=row_idx, column=5, value=round(candidate.fit_score, 2))
                ws2.cell(row=row_idx, column=6, value='Yes' if candidate.certification_match.get('has_must_have') else 'No')
                ws2.cell(row=row_idx, column=7, value='Yes' if candidate.certification_match.get('has_bonus') else 'No')
                ws2.cell(row=row_idx, column=8, value=f"{candidate.skills_match.get('required_match_rate', 0.0):.2%}" if candidate.skills_match else 'N/A')
                ws2.cell(row=row_idx, column=9, value=f"{candidate.skills_match.get('preferred_match_rate', 0.0):.2%}" if candidate.skills_match else 'N/A')
                ws2.cell(row=row_idx, column=10, value=f"{candidate.experience_match.get('level_match', 0.0):.2%}" if candidate.experience_match else 'N/A')
                ws2.cell(row=row_idx, column=11, value='Yes' if candidate.location_match else 'No')
                ws2.cell(row=row_idx, column=12, value='; '.join(candidate.certifications) if candidate.certifications else 'None')
                ws2.cell(row=row_idx, column=13, value=candidate.experience_match.get('years', 0) if candidate.experience_match else 0)
                ws2.cell(row=row_idx, column=14, value=candidate.rationale[:500] if candidate.rationale else '')
            
            # Auto-adjust column widths
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError, ValueError) as e:
                            logger.debug(f"Error processing cell value: {e}")
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
    except Exception as e:
        logger.error(f"Error generating Excel export: {e}", exc_info=True)
        return None


def generate_json_export(analysis_data: Dict, job_title: str, location: str) -> bytes:
    """Generate JSON export with full analysis data"""
    export_data = {
        'job_title': job_title,
        'location': location,
        'export_date': datetime.now().isoformat(),
        'candidates': []
    }
    
    for candidate in sorted(analysis_data.get('candidate_scores', []), 
                           key=lambda x: x.fit_score, reverse=True):
        export_data['candidates'].append({
            'name': candidate.name,
            'email': candidate.email,
            'phone': candidate.phone,
            'fit_score': candidate.fit_score,
            'certifications': candidate.certifications,
            'certification_match': candidate.certification_match,
            'skills_match': candidate.skills_match,
            'experience_match': candidate.experience_match,
            'location_match': candidate.location_match,
            'component_scores': candidate.component_scores,
            'rationale': candidate.rationale,
            'chain_of_thought': candidate.chain_of_thought
        })
    
    json_str = json.dumps(export_data, indent=2, default=str)
    return json_str.encode('utf-8')


def _make_download_link(data_bytes: bytes, filename: str, mime: str) -> str:
    """Return an HTML download link using base64 to avoid transient media cache misses."""
    if not data_bytes:
        return ""
    b64 = base64.b64encode(data_bytes).decode()
    return f'<a class="download-link" href="data:{mime};base64,{b64}" download="{filename}"><span style="font-size: 1.2em;">↓</span> Download {filename}</a>'


def get_score_color(score):
    """Return color class based on score"""
    if score >= 7.5:
        return "#27ae60"  # Green - Excellent
    elif score >= 5.5:
        return "#f39c12"  # Yellow/Orange - Good
    else:
        return "#e74c3c"  # Red - Needs improvement


def display_analysis_history(user_id: str):
    """
    Display list of user's previous analyses with search, filter, sort, and management options.
    """
    logger.info(f"Displaying analysis history for user {user_id}")
    
    try:
        st.markdown('<div class="section-header">Previous Analyses</div>', unsafe_allow_html=True)
        
        # Validate user_id
        if not user_id:
            st.error("**Error:** User ID is missing. Please log in again.")
            logger.error("display_analysis_history called with empty user_id")
            return
        
        # Load analyses with error handling
        try:
            analyses = load_user_analyses(user_id)
            logger.info(f"Loaded {len(analyses)} analyses for user {user_id}")
        except Exception as e:
            logger.error(f"Error loading analyses for user {user_id}: {e}", exc_info=True)
            st.error("**Error:** Error loading analyses. Please try again or contact support if the issue persists.")
            return
        
        if not analyses:
            st.info("You haven't run any analyses yet. Start a new analysis to see results here.")
            return
        
        total_count = len(analyses)
        
        # Search and Filter Section
        with st.expander("Search & Filter", expanded=False):
            col_search, col_date1, col_date2 = st.columns([2, 1, 1])
        
        with col_search:
            search_query = st.text_input(
                "Search by job title",
                value=st.session_state.get("analysis_search", ""),
                key="analysis_search_input",
                placeholder="Enter job title to search..."
            )
            st.session_state["analysis_search"] = search_query
        
        with col_date1:
            # Use session state default to avoid modification after widget instantiation
            date_from = st.date_input(
                "From Date",
                value=st.session_state.get("filter_date_from", None),
                key="filter_date_from"
            )
        
        with col_date2:
            # Use session state default to avoid modification after widget instantiation
            date_to = st.date_input(
                "To Date",
                value=st.session_state.get("filter_date_to", None),
                key="filter_date_to"
            )
        
        col_loc, col_cand1, col_cand2 = st.columns([2, 1, 1])
        
        with col_loc:
            # Get unique locations
            unique_locations = sorted(list(set([a['location'] for a in analyses if a.get('location')])))
            location_filter = st.selectbox(
                "Filter by Location",
                options=["All"] + unique_locations,
                key="filter_location"
            )
        
        with col_cand1:
            min_candidates = st.number_input(
                "Min Candidates",
                min_value=0,
                max_value=1000,
                value=0,
                key="filter_min_candidates"
            )
        
        with col_cand2:
            max_candidates = st.number_input(
                "Max Candidates",
                min_value=0,
                max_value=1000,
                value=1000,
                key="filter_max_candidates"
            )
        
        col_clear, col_sort = st.columns([1, 2])
        with col_clear:
            if st.button("Clear Filters", key="clear_filters"):
                # Clear filters by rerunning with cleared query params
                # Don't modify session state directly for widgets
                st.query_params.clear()
                st.rerun()
        
        with col_sort:
            sort_option = st.selectbox(
                "Sort By",
                options=[
                    "Date (Newest First)",
                    "Date (Oldest First)",
                    "Job Title (A-Z)",
                    "Job Title (Z-A)",
                    "Candidates (Most First)",
                    "Candidates (Least First)"
                ],
                key="analysis_sort"
            )
        
        # Apply filters
        filtered_analyses = analyses.copy()
        
        # Search filter
        if search_query:
            search_lower = search_query.lower()
            filtered_analyses = [
                a for a in filtered_analyses
                if search_lower in a.get('title', '').lower()
            ]
        
        # Date filter
        if date_from or date_to:
            filtered_analyses = [
                a for a in filtered_analyses
                if _matches_date_filter(a.get('created_at', ''), date_from, date_to)
            ]
        
        # Location filter
        if location_filter != "All":
            filtered_analyses = [
                a for a in filtered_analyses
                if a.get('location') == location_filter
            ]
        
        # Candidate count filter
        filtered_analyses = [
            a for a in filtered_analyses
            if min_candidates <= a.get('num_candidates', 0) <= max_candidates
        ]
        
        # Apply sorting
        filtered_analyses = _sort_analyses(filtered_analyses, sort_option)
        
        # Display filtered count
        if len(filtered_analyses) != total_count:
            st.info(f"Showing {len(filtered_analyses)} of {total_count} analyses")
        else:
            st.info(f"Found {total_count} previous analysis(es)")
        
        if not filtered_analyses:
            st.warning("No analyses match your search and filter criteria.")
            return
        
        # Pagination
        items_per_page = st.session_state.get("analyses_per_page", 10)
        page_num = st.session_state.get("analysis_page", 0)
        
        col_page1, col_page2, col_page3 = st.columns([1, 2, 1])
        with col_page2:
            items_per_page = st.selectbox(
                "Items per page",
                options=[10, 20, 50, 100],
                index=[10, 20, 50, 100].index(items_per_page) if items_per_page in [10, 20, 50, 100] else 0,
                key="analyses_per_page_select"
            )
            st.session_state["analyses_per_page"] = items_per_page
        
        total_pages = (len(filtered_analyses) + items_per_page - 1) // items_per_page
        start_idx = page_num * items_per_page
        end_idx = min(start_idx + items_per_page, len(filtered_analyses))
        paginated_analyses = filtered_analyses[start_idx:end_idx]
        
        # Page navigation
        if total_pages > 1:
            nav_col1, nav_col2, nav_col3, nav_col4, nav_col5 = st.columns([1, 1, 2, 1, 1])
            with nav_col1:
                if st.button("◀◀ First", key="page_first", disabled=(page_num == 0)):
                    st.session_state["analysis_page"] = 0
                    st.rerun()
            with nav_col2:
                if st.button("◀ Prev", key="page_prev", disabled=(page_num == 0)):
                    st.session_state["analysis_page"] = max(0, page_num - 1)
                    st.rerun()
            with nav_col3:
                st.markdown(f"<div style='text-align: center; padding-top: 0.5rem;'>Page {page_num + 1} of {total_pages}</div>", unsafe_allow_html=True)
            with nav_col4:
                if st.button("Next ▶", key="page_next", disabled=(page_num >= total_pages - 1)):
                    st.session_state["analysis_page"] = min(total_pages - 1, page_num + 1)
                    st.rerun()
            with nav_col5:
                if st.button("Last ▶▶", key="page_last", disabled=(page_num >= total_pages - 1)):
                    st.session_state["analysis_page"] = total_pages - 1
                    st.rerun()
        
        # Display analyses in card format
        for idx, analysis in enumerate(paginated_analyses):
            # Parse date for color coding
            try:
                created_at_str = analysis['created_at']
                if created_at_str.endswith('Z'):
                    dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                else:
                    dt = datetime.fromisoformat(created_at_str)
                analysis_date = dt.date()
                date_str = dt.strftime('%Y-%m-%d')
                time_str = dt.strftime('%H:%M:%S')
                days_ago = (datetime.now().date() - analysis_date).days
            except (ValueError, AttributeError, TypeError):
                date_str = analysis['created_at'][:10] if len(analysis['created_at']) >= 10 else 'Unknown'
                time_str = analysis['created_at'][11:19] if len(analysis['created_at']) > 19 else ''
                days_ago = 999
            
            # Card styling based on age
            if days_ago <= 7:
                border_color = "#00A651"  # Green for recent
            elif days_ago <= 30:
                border_color = "#0066CC"  # Blue for recent month
            else:
                border_color = "#6c757d"  # Gray for older (adjusted for dark mode visibility)
            
            # Card container
            with st.container():
                st.markdown(
                    f"""
                    <div style="
                        border: 2px solid {border_color};
                        border-radius: 0.5rem;
                        padding: var(--spacing-sm);
                        margin-bottom: var(--spacing-sm);
                        background: linear-gradient(135deg, var(--bg-primary) 0%, var(--bg-secondary) 100%);
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        color: var(--text-primary);
                    ">
                    """,
                    unsafe_allow_html=True
                )
                
                # Card header with title and actions
                col_header1, col_header2 = st.columns([3, 1])
                with col_header1:
                    st.markdown(f"### {analysis['title']}")
                    st.caption(f"Location: {analysis['location']}")
                with col_header2:
                    # Quick stats
                    st.metric("Candidates", analysis['num_candidates'])
                
                # Card body with details
                col_body1, col_body2 = st.columns([2, 1])
                with col_body1:
                    st.markdown(f"**Date:** {date_str} at {time_str}")
                    if days_ago == 0:
                        st.caption("Created today")
                    elif days_ago <= 7:
                        st.caption(f"Created {days_ago} day(s) ago")
                    elif days_ago <= 30:
                        st.caption(f"Created {days_ago} day(s) ago")
                    else:
                        st.caption(f"Created {days_ago} day(s) ago")
                
                with col_body2:
                    # Preview expander
                    with st.expander("Quick Preview", expanded=False):
                        analysis_data = load_analysis_data(analysis['report_id'], user_id)
                        if analysis_data and analysis_data.get('candidate_scores'):
                            scores = [cs.fit_score for cs in analysis_data['candidate_scores']]
                            if scores:
                                avg_score = sum(scores) / len(scores)
                                max_score = max(scores)
                                min_score = min(scores)
                                st.metric("Avg Score", f"{avg_score:.2f}")
                                st.caption(f"Range: {min_score:.2f} - {max_score:.2f}")
                                # Top 3 candidates
                                top_3 = sorted(analysis_data['candidate_scores'], key=lambda x: x.fit_score, reverse=True)[:3]
                                st.markdown("**Top 3:**")
                                for i, cand in enumerate(top_3, 1):
                                    st.caption(f"{i}. {cand.name}: {cand.fit_score:.2f}/10")
                        else:
                            st.caption("Preview unavailable")
                
                # Action buttons
                col_actions1, col_actions2, col_actions3, col_actions4, col_actions5 = st.columns([1, 1, 1, 1, 1])
                
                with col_actions1:
                    if st.button("View", key=f"view_{analysis['report_id']}", width='stretch'):
                        st.session_state["view_report_id"] = analysis['report_id']
                        st.session_state["viewing_previous_analysis"] = True
                        st.rerun()
                
                with col_actions2:
                    # PDF download
                    pdf_data = None
                    if analysis['pdf_path'] and os.path.exists(analysis['pdf_path']):
                        try:
                            if is_safe_path(analysis['pdf_path']):
                                file_size = os.path.getsize(analysis['pdf_path'])
                                if file_size <= MAX_PDF_SIZE:
                                    with open(analysis['pdf_path'], 'rb') as f:
                                        pdf_data = f.read()
                        except Exception:
                            pass
                    
                    if pdf_data:
                        pdf_link = _make_download_link(
                            pdf_data,
                            os.path.basename(analysis['pdf_path']),
                            "application/pdf"
                        )
                        if pdf_link:
                            st.markdown(pdf_link, unsafe_allow_html=True)
                    else:
                        st.caption("PDF N/A")
                
                with col_actions3:
                    # CSV download
                    analysis_data = load_analysis_data(analysis['report_id'], user_id)
                    if analysis_data and analysis_data.get('candidate_scores'):
                        csv_data = generate_csv_export(
                            analysis_data['candidate_scores'],
                            analysis['title'],
                            analysis['location']
                        )
                        csv_link = _make_download_link(
                            csv_data.encode("utf-8"),
                            f"candidates_{analysis['created_at'][:10].replace('-', '')}.csv",
                            "text/csv"
                        )
                        if csv_link:
                            st.markdown(csv_link, unsafe_allow_html=True)
                        else:
                            st.caption("CSV N/A")
                
                with col_actions4:
                    # Excel and JSON downloads
                    analysis_data = load_analysis_data(analysis['report_id'], user_id)
                    if analysis_data and analysis_data.get('candidate_scores'):
                        col_excel, col_json = st.columns(2)
                        with col_excel:
                            excel_data = generate_excel_export(
                                analysis_data,
                                analysis['title'],
                                analysis['location']
                            )
                            if excel_data:
                                excel_link = _make_download_link(
                                    excel_data,
                                    f"analysis_{analysis['created_at'][:10].replace('-', '')}.xlsx",
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                if excel_link:
                                    st.markdown(excel_link, unsafe_allow_html=True)
                            else:
                                st.caption("Excel N/A")
                        with col_json:
                            json_data = generate_json_export(
                                analysis_data,
                                analysis['title'],
                                analysis['location']
                            )
                            json_link = _make_download_link(
                                json_data,
                                f"analysis_{analysis['created_at'][:10].replace('-', '')}.json",
                                "application/json"
                            )
                            if json_link:
                                st.markdown(json_link, unsafe_allow_html=True)
                    else:
                        st.caption("Export N/A")
                
                with col_actions5:
                    # Delete button
                    delete_key = f"delete_{analysis['report_id']}"
                    if delete_key not in st.session_state:
                        st.session_state[delete_key] = False
                    
                    if st.button("Delete", key=f"del_btn_{analysis['report_id']}", width='stretch', type="secondary"):
                        st.session_state[delete_key] = True
                        st.rerun()
                    
                    if st.session_state.get(delete_key):
                        st.warning(f"Delete '{analysis['title']}'?")
                        col_confirm1, col_confirm2 = st.columns(2)
                        with col_confirm1:
                            if st.button("Confirm", key=f"confirm_del_{analysis['report_id']}"):
                                if delete_analysis(analysis['report_id'], user_id):
                                    st.success("Analysis deleted successfully!")
                                    st.session_state.pop(delete_key, None)
                                    # Reset pagination if needed
                                    if page_num > 0 and len(filtered_analyses) <= page_num * items_per_page:
                                        st.session_state["analysis_page"] = max(0, page_num - 1)
                                    st.rerun()
                                else:
                                    st.error("Failed to delete analysis.")
                                    st.session_state.pop(delete_key, None)
                        with col_confirm2:
                            if st.button("Cancel", key=f"cancel_del_{analysis['report_id']}"):
                                st.session_state.pop(delete_key, None)
                                st.rerun()
                
                st.markdown("</div>", unsafe_allow_html=True)
                
                if idx < len(paginated_analyses) - 1:
                    st.markdown("<br>", unsafe_allow_html=True)
    
    except Exception as e:
        logger.error(f"Error in display_analysis_history for user {user_id}: {e}", exc_info=True)
        st.error("**Error:** An unexpected error occurred while displaying analysis history. Please try again or contact support if the issue persists.")
        st.exception(e)


def display_analysis_analytics(user_id: str):
    """Display analytics dashboard for user's analyses"""
    logger.info(f"Displaying analytics for user {user_id}")
    
    try:
        st.markdown('<div class="section-header">Analysis Analytics</div>', unsafe_allow_html=True)
        
        # Load analyses with error handling
        try:
            analyses = load_user_analyses(user_id)
            logger.info(f"Loaded {len(analyses)} analyses for user {user_id}")
        except Exception as e:
            logger.error(f"Error loading analyses for user {user_id}: {e}", exc_info=True)
            st.error("**Error:** Error loading analyses. Please try again or contact support if the issue persists.")
            return
        
        if not analyses:
            st.info("No analyses available for analytics. Run some analyses first!")
            return
        
        # Calculate statistics
        total_analyses = len(analyses)
        total_candidates = sum(a.get('num_candidates', 0) for a in analyses)
        avg_candidates = total_candidates / total_analyses if total_analyses > 0 else 0
        
        # Date range
        dates = []
        for a in analyses:
            try:
                created_at_str = a.get('created_at', '')
                if created_at_str.endswith('Z'):
                    dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                else:
                    dt = datetime.fromisoformat(created_at_str)
                dates.append(dt.date())
            except Exception as e:
                logger.debug(f"Error parsing date: {e}")
                pass
        
        first_date = min(dates) if dates else None
        last_date = max(dates) if dates else None
        
        # Most common job title
        job_titles = [a.get('title', '') for a in analyses if a.get('title')]
        most_common_title = max(set(job_titles), key=job_titles.count) if job_titles else 'N/A'
        
        # Most common location
        locations = [a.get('location', '') for a in analyses if a.get('location')]
        most_common_location = max(set(locations), key=locations.count) if locations else 'N/A'
        
        # Display key metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Analyses", total_analyses)
        with col2:
            st.metric("Total Candidates", total_candidates)
        with col3:
            st.metric("Avg Candidates/Analysis", f"{avg_candidates:.1f}")
        with col4:
            st.metric("Date Range", f"{first_date} to {last_date}" if first_date and last_date else "N/A")
        
        st.markdown("---")
        
        # Check if pandas is available
        try:
            import pandas as pd
            pandas_available = True
        except ModuleNotFoundError:
            pandas_available = False
            logger.warning("pandas module not found - charts will be limited")
            st.warning("**Warning:** Charts require pandas. Install with: `pip install pandas`")
        
        # Charts section
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            st.markdown(f"<h3 style='font-size: var(--font-size-xl); font-weight: 600; margin-bottom: var(--spacing-sm);'>Analyses Over Time</h3>", unsafe_allow_html=True)
            # Group by month
            monthly_counts = {}
            for a in analyses:
                try:
                    created_at_str = a.get('created_at', '')
                    if created_at_str.endswith('Z'):
                        dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        dt = datetime.fromisoformat(created_at_str)
                    month_key = dt.strftime('%Y-%m')
                    monthly_counts[month_key] = monthly_counts.get(month_key, 0) + 1
                except Exception as e:
                    logger.debug(f"Error parsing date for analysis: {e}")
                    pass
            
            if monthly_counts:
                if pandas_available:
                    try:
                        chart_data = pd.DataFrame({
                            'Month': list(monthly_counts.keys()),
                            'Count': list(monthly_counts.values())
                        }).sort_values('Month')
                        st.bar_chart(chart_data.set_index('Month'))
                    except Exception as e:
                        logger.error(f"Error creating monthly chart: {e}", exc_info=True)
                        st.error("Error displaying chart. Showing data in table format instead.")
                        # Fallback: show as table
                        for month, count in sorted(monthly_counts.items()):
                            st.text(f"{month}: {count} analyses")
                else:
                    # Fallback without pandas
                    st.info("Monthly breakdown:")
                    for month, count in sorted(monthly_counts.items()):
                        st.text(f"{month}: {count} analyses")
        
        with col_chart2:
            st.markdown(f"<h3 style='font-size: var(--font-size-xl); font-weight: 600; margin-bottom: var(--spacing-sm);'>Job Titles Distribution</h3>", unsafe_allow_html=True)
            # Count job titles
            title_counts = {}
            for title in job_titles:
                title_counts[title] = title_counts.get(title, 0) + 1
            
            if title_counts:
                # Show top 10
                top_titles = sorted(title_counts.items(), key=lambda x: x[1], reverse=True)[:10]
                if pandas_available:
                    try:
                        chart_data = pd.DataFrame({
                            'Job Title': [t[0][:30] for t in top_titles],  # Truncate long titles
                            'Count': [t[1] for t in top_titles]
                        })
                        st.bar_chart(chart_data.set_index('Job Title'))
                    except Exception as e:
                        logger.error(f"Error creating job titles chart: {e}", exc_info=True)
                        st.error("Error displaying chart. Showing data in table format instead.")
                        # Fallback: show as list
                        for title, count in top_titles:
                            st.text(f"{title}: {count} analyses")
                else:
                    # Fallback without pandas
                    st.info("Top job titles:")
                    for title, count in top_titles:
                        st.text(f"{title}: {count} analyses")
        
        # Additional statistics
        st.markdown("---")
        st.markdown(f"<h3 style='font-size: var(--font-size-xl); font-weight: 600; margin-bottom: var(--spacing-sm);'>Summary Statistics</h3>", unsafe_allow_html=True)
        
        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.markdown(f"**Most Analyzed Job Title:** {most_common_title}")
            st.markdown(f"**Most Common Location:** {most_common_location}")
        with col_stat2:
            st.markdown(f"**First Analysis:** {first_date}" if first_date else "**First Analysis:** N/A")
            st.markdown(f"**Latest Analysis:** {last_date}" if last_date else "**Latest Analysis:** N/A")
        
        # Candidate count trend
        if len(analyses) > 1:
            st.markdown("---")
            st.markdown(f"<h3 style='font-size: var(--font-size-xl); font-weight: 600; margin-bottom: var(--spacing-sm);'>Candidate Count Trend</h3>", unsafe_allow_html=True)
            candidate_trend = []
            for a in sorted(analyses, key=lambda x: x.get('created_at', '')):
                try:
                    created_at_str = a.get('created_at', '')
                    if created_at_str.endswith('Z'):
                        dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                    else:
                        dt = datetime.fromisoformat(created_at_str)
                    candidate_trend.append({
                        'Date': dt.strftime('%Y-%m-%d'),
                        'Candidates': a.get('num_candidates', 0)
                    })
                except Exception as e:
                    logger.debug(f"Error parsing date for trend: {e}")
                    pass
            
            if candidate_trend:
                if pandas_available:
                    try:
                        trend_df = pd.DataFrame(candidate_trend)
                        trend_df['Date'] = pd.to_datetime(trend_df['Date'])
                        trend_df = trend_df.sort_values('Date')
                        st.line_chart(trend_df.set_index('Date'))
                    except Exception as e:
                        logger.error(f"Error creating trend chart: {e}", exc_info=True)
                        st.error("Error displaying trend chart. Showing data in table format instead.")
                        # Fallback: show as table
                        for item in sorted(candidate_trend, key=lambda x: x['Date']):
                            st.text(f"{item['Date']}: {item['Candidates']} candidates")
                else:
                    # Fallback without pandas
                    st.info("Candidate count trend:")
                    for item in sorted(candidate_trend, key=lambda x: x['Date']):
                        st.text(f"{item['Date']}: {item['Candidates']} candidates")
    
    except Exception as e:
        logger.error(f"Error in display_analysis_analytics for user {user_id}: {e}", exc_info=True)
        st.error("**Error:** An unexpected error occurred while displaying analytics. Please try again or contact support if the issue persists.")
        st.exception(e)


def display_resume_database(user_id: str):
    """Display resume database with search, filter, and management options"""
    logger.info(f"Displaying resume database for user {user_id}")
    
    try:
        st.markdown('<div class="section-header">Resume Database</div>', unsafe_allow_html=True)
        
        # Validate user_id
        if not user_id:
            st.error("**Error:** User ID is missing. Please log in again.")
            logger.error("display_resume_database called with empty user_id")
            return
        
        # Search and Filter Section
        with st.expander("Search & Filter", expanded=True):
            col1, col2 = st.columns([2, 1])
            
            with col1:
                search_query = st.text_input(
                    "Search by name, email, skills, or certifications",
                    key="resume_search",
                    placeholder="Enter search terms..."
                )
            
            with col2:
                experience_filter = st.selectbox(
                    "Experience Level",
                    ["All", "Junior", "Mid", "Senior"],
                    key="resume_exp_filter"
                )
            
            col3, col4 = st.columns(2)
            with col3:
                location_filter = st.text_input(
                    "Location",
                    key="resume_location_filter",
                    placeholder="Filter by location..."
                )
            
            with col4:
                tag_filter = st.text_input(
                    "Tags (comma-separated)",
                    key="resume_tag_filter",
                    placeholder="e.g., safety, certified, experienced"
                )
        
        # Build filters
        filters = {}
        if experience_filter != "All":
            filters['experience_level'] = experience_filter
        if location_filter:
            filters['location'] = location_filter
        if tag_filter:
            filters['tags'] = [t.strip() for t in tag_filter.split(',') if t.strip()]
        
        # Search candidates with error handling
        try:
            candidates = search_candidates(user_id, query=search_query or "", filters=filters)
            logger.info(f"Found {len(candidates)} candidates for user {user_id}")
        except Exception as e:
            logger.error(f"Error searching candidates for user {user_id}: {e}", exc_info=True)
            st.error("**Error:** Error searching candidates. Please try again or contact support if the issue persists.")
            candidates = []
        
        # Display candidate count
        st.markdown(f"**Found {len(candidates)} candidate(s)**")
        
        if not candidates:
            st.info("No candidates found. Upload resumes to build your database!")
            
            # Upload new resume section
            st.markdown("---")
            st.markdown("### Upload New Resume")
            uploaded_file = st.file_uploader(
                "Upload a resume to add to your database",
                type=["pdf", "docx", "txt"],
                key="resume_db_upload"
            )
            
            if uploaded_file:
                if st.button("Save to Database", type="primary"):
                    try:
                        # Parse the resume
                        from resume_parser import ResumeParser
                        import tempfile
                        import shutil
                        
                        temp_dir = tempfile.mkdtemp()
                        temp_path = os.path.join(temp_dir, uploaded_file.name)
                        
                        try:
                            with open(temp_path, 'wb') as f:
                                f.write(uploaded_file.getbuffer())
                            
                            parser = ResumeParser()
                            resume_data = parser.parse(temp_path)
                            
                            # Save file
                            file_bytes = uploaded_file.getbuffer()
                            stored_path, file_hash = save_bytes(file_bytes, uploaded_file.name)
                            
                            # Save to file_assets
                            import uuid
                            asset_id = str(uuid.uuid4())
                            
                            with get_db() as conn:
                                cur = conn.cursor()
                                query = prepare_query(conn, """
                                    INSERT INTO file_assets 
                                    (id, user_id, kind, original_name, stored_path, created_at)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """)
                                cur.execute(query, (
                                    asset_id,
                                    user_id,
                                    'resume',
                                    uploaded_file.name,
                                    stored_path,
                                    utcnow_str()
                                ))
                                conn.commit()
                            
                            # Save to candidate profiles
                            profile_id = save_candidate_profile(
                                user_id=user_id,
                                resume_data=resume_data,
                                resume_file_id=asset_id,
                                tags=[],
                                notes=""
                            )
                            
                            st.success(f"Resume saved to database! Profile ID: {profile_id}")
                            logger.info(f"Resume saved successfully for user {user_id}, profile_id: {profile_id}")
                            st.rerun()
                        finally:
                            try:
                                shutil.rmtree(temp_dir)
                            except Exception as e:
                                logger.warning(f"Failed to cleanup temp directory: {e}")
                    except Exception as e:
                        logger.error(f"Error saving resume to database for user {user_id}: {e}", exc_info=True)
                        st.error(f"**Error:** Error saving resume: {str(e)}")
            return
    
        # Display candidates in a grid/list
        for idx, candidate in enumerate(candidates):
            with st.container():
                col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
                
                with col1:
                    st.markdown(f"**{candidate['name'] or 'Unknown'}**")
                    if candidate['email']:
                        st.caption(f"📧 {candidate['email']}")
                    if candidate['phone']:
                        st.caption(f"📞 {candidate['phone']}")
                    if candidate['location']:
                        st.caption(f"📍 {candidate['location']}")
                
                with col2:
                    parsed_data = candidate.get('parsed_data', {})
                    skills = parsed_data.get('skills', [])[:5]
                    if skills:
                        st.markdown("**Skills:**")
                        st.caption(", ".join(skills))
                    else:
                        st.caption("No skills listed")
                
                with col3:
                    certs = parsed_data.get('certifications', [])[:3]
                    if certs:
                        st.markdown("**Certifications:**")
                        st.caption(", ".join(certs))
                    else:
                        st.caption("No certifications")
                    
                    # Tags
                    tags = candidate.get('tags', [])
                    if tags:
                        tag_text = " ".join([f"`{tag}`" for tag in tags])
                        st.caption(f"Tags: {tag_text}")
                
                with col4:
                    # Actions
                    action_key = f"action_{candidate['id']}_{idx}"
                    action = st.selectbox(
                        "Actions",
                        ["Select...", "View", "Edit", "Delete", "Use in Analysis"],
                        key=action_key
                    )
                    
                    if action == "View":
                        if st.session_state.get(f"view_profile_{candidate['id']}") != True:
                            st.session_state[f"view_profile_{candidate['id']}"] = True
                            st.rerun()
                    
                    if action == "Delete":
                        if st.button("Confirm Delete", key=f"delete_{candidate['id']}"):
                            if delete_candidate_profile(candidate['id'], user_id):
                                st.success("Candidate deleted!")
                                st.rerun()
                            else:
                                st.error("Failed to delete candidate")
                    
                    if action == "Use in Analysis":
                        # Store selected candidates in session state
                        if 'selected_candidates' not in st.session_state:
                            st.session_state['selected_candidates'] = []
                        if candidate['id'] not in st.session_state['selected_candidates']:
                            st.session_state['selected_candidates'].append(candidate['id'])
                            st.success("Added to selection!")
                
                # View profile details if requested
                if st.session_state.get(f"view_profile_{candidate['id']}"):
                    with st.expander(f"View Profile: {candidate['name']}", expanded=True):
                        st.markdown(f"**Name:** {candidate['name']}")
                        st.markdown(f"**Email:** {candidate['email']}")
                        st.markdown(f"**Phone:** {candidate['phone']}")
                        st.markdown(f"**Location:** {candidate['location']}")
                        
                        parsed_data = candidate.get('parsed_data', {})
                        if parsed_data.get('skills'):
                            st.markdown(f"**Skills:** {', '.join(parsed_data['skills'])}")
                        if parsed_data.get('certifications'):
                            st.markdown(f"**Certifications:** {', '.join(parsed_data['certifications'])}")
                        if parsed_data.get('job_titles'):
                            st.markdown(f"**Job Titles:** {', '.join(parsed_data['job_titles'])}")
                        if parsed_data.get('years_of_experience'):
                            st.markdown(f"**Years of Experience:** {parsed_data['years_of_experience']}")
                        
                        if candidate.get('notes'):
                            st.markdown(f"**Notes:** {candidate['notes']}")
                        
                        # Edit form
                        with st.form(f"edit_form_{candidate['id']}"):
                            new_tags = st.text_input(
                                "Tags (comma-separated)",
                                value=", ".join(candidate.get('tags', [])),
                                key=f"tags_{candidate['id']}"
                            )
                            new_notes = st.text_area(
                                "Notes",
                                value=candidate.get('notes', ''),
                                key=f"notes_{candidate['id']}"
                            )
                            
                            if st.form_submit_button("Save Changes"):
                                updates = {}
                                if new_tags:
                                    updates['tags'] = [t.strip() for t in new_tags.split(',') if t.strip()]
                                if new_notes is not None:
                                    updates['notes'] = new_notes
                                
                                if updates:
                                    if update_candidate_profile(candidate['id'], user_id, updates):
                                        st.success("Profile updated!")
                                        st.session_state.pop(f"view_profile_{candidate['id']}", None)
                                        st.rerun()
                                    else:
                                        st.error("Failed to update profile")
                        
                        if st.button("Close", key=f"close_{candidate['id']}"):
                            st.session_state.pop(f"view_profile_{candidate['id']}", None)
                            st.rerun()
                
                if idx < len(candidates) - 1:
                    st.markdown("---")
        
        # Upload new resume section at bottom
        st.markdown("---")
        st.markdown("### Upload New Resume")
        uploaded_file = st.file_uploader(
            "Upload a resume to add to your database",
            type=["pdf", "docx", "txt"],
            key="resume_db_upload_bottom"
        )
        
        if uploaded_file:
            tags_input = st.text_input("Tags (comma-separated)", key="new_resume_tags")
            notes_input = st.text_area("Notes", key="new_resume_notes")
            
            if st.button("Save to Database", type="primary", key="save_new_resume"):
                try:
                    # Parse the resume
                    from resume_parser import ResumeParser
                    import tempfile
                    import shutil
                    
                    temp_dir = tempfile.mkdtemp()
                    temp_path = os.path.join(temp_dir, uploaded_file.name)
                    
                    try:
                        with open(temp_path, 'wb') as f:
                            f.write(uploaded_file.getbuffer())
                        
                        parser = ResumeParser()
                        resume_data = parser.parse(temp_path)
                        
                        # Save file
                        file_bytes = uploaded_file.getbuffer()
                        stored_path, file_hash = save_bytes(file_bytes, uploaded_file.name)
                        
                        # Save to file_assets
                        import uuid
                        asset_id = str(uuid.uuid4())
                        
                        with get_db() as conn:
                            cur = conn.cursor()
                            query = prepare_query(conn, """
                                INSERT INTO file_assets 
                                (id, user_id, kind, original_name, stored_path, created_at)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """)
                            cur.execute(query, (
                                asset_id,
                                user_id,
                                'resume',
                                uploaded_file.name,
                                stored_path,
                                utcnow_str()
                            ))
                            try:
                                conn.commit()
                            except Exception as e:
                                conn.rollback()
                                logger.error(f"Failed to save file asset: {e}", exc_info=True)
                                raise
                        
                        # Parse tags
                        tags = [t.strip() for t in tags_input.split(',') if t.strip()] if tags_input else []
                        
                        # Save to candidate profiles
                        profile_id = save_candidate_profile(
                            user_id=user_id,
                            resume_data=resume_data,
                            resume_file_id=asset_id,
                            tags=tags,
                            notes=notes_input or ""
                        )
                        
                        st.success(f"Resume saved to database!")
                        logger.info(f"Resume saved successfully for user {user_id}, profile_id: {profile_id}")
                        st.rerun()
                    finally:
                        try:
                            shutil.rmtree(temp_dir)
                        except Exception as e:
                            logger.warning(f"Failed to cleanup temp directory: {e}")
                except Exception as e:
                    logger.error(f"Error saving resume to database for user {user_id}: {e}", exc_info=True)
                    st.error(f"**Error:** Error saving resume: {str(e)}")
    
    except Exception as e:
        logger.error(f"Error in display_resume_database for user {user_id}: {e}", exc_info=True)
        st.error("**Error:** An unexpected error occurred while displaying the resume database. Please try again or contact support if the issue persists.")
        st.exception(e)


def display_results(results):
    """Display stored results with sorting and filtering options"""
    st.markdown('<div class="section-header">Results</div>', unsafe_allow_html=True)

    # Lightweight styling for download anchors
    st.markdown("""
        <style>
        a.download-link {
            display: inline-block;
            padding: 0.75rem 1rem;
            border-radius: 0.5rem;
            background: var(--brand-blue);
            color: white !important;
            text-decoration: none;
            font-weight: 600;
            width: 100%;
            text-align: center;
        }
        a.download-link:hover { background: var(--info-border); }
        </style>
    """, unsafe_allow_html=True)

    candidate_scores = results['candidate_scores']
    viable_candidates = [c for c in candidate_scores if c.fit_score >= 5.0]
    pdf_path = results['pdf_path']
    pdf_data = results['pdf_data']
    job_details = results['job_details']
    processing_time = results.get('processing_time', 0)

    # Results header with navigation buttons
    col1, col2, col3 = st.columns([3, 1, 1])
    with col1:
        st.markdown(f"**Job:** {job_details['title']} | **Location:** {job_details['location']}")
    with col2:
        if st.button("Back to Analyses", type="secondary", help="Return to Previous Analyses list"):
            st.session_state["show_history_list"] = True
            st.session_state.pop("viewing_previous_analysis", None)
            st.rerun()
    with col3:
        if st.button("Start New Analysis", type="secondary", help="Start a new analysis"):
            clear_results()
            st.session_state.pop("viewing_previous_analysis", None)
            st.rerun()

    # Summary metrics with color coding
    st.markdown("### Summary")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Candidates", len(candidate_scores))

    with col2:
        viable_count = len(viable_candidates)
        st.metric("Viable Candidates (>=5.0)", viable_count)

    with col3:
        if viable_candidates:
            avg_score = sum(c.fit_score for c in viable_candidates) / len(viable_candidates)
            st.metric("Average Score (Viable)", f"{avg_score:.2f}/10")
        else:
            st.metric("Average Score (Viable)", "N/A")

    with col4:
        if processing_time > 0:
            st.metric("Processing Time", f"{processing_time:.1f}s")

    # Ranking Visualization
    if viable_candidates:
        st.markdown("### Ranking Visualization")
        
        # Sort candidates by score (highest to lowest) for visualization
        sorted_for_viz = sorted(viable_candidates, key=lambda x: x.fit_score, reverse=True)
        
        try:
            import pandas as pd
            
            # Create DataFrame for bar chart
            chart_data = pd.DataFrame({
                'Candidate': [c.name[:30] + ('...' if len(c.name) > 30 else '') for c in sorted_for_viz],
                'Score': [c.fit_score for c in sorted_for_viz]
            })
            
            # Create color mapping based on score (using CSS variables)
            # Note: Plotly doesn't support CSS variables directly, so we use the actual color values
            from ui.theme import SCORE_COLORS
            colors = []
            for score in chart_data['Score']:
                if score >= 7.5:
                    colors.append(SCORE_COLORS['excellent'])  # Green - Excellent
                elif score >= 5.5:
                    colors.append(SCORE_COLORS['good'])  # Yellow/Orange - Good
                else:
                    colors.append(SCORE_COLORS['poor'])  # Red - Needs improvement
            
            # Display horizontal bar chart
            st.bar_chart(chart_data.set_index('Candidate'), height=400, use_container_width=True)
            
            # Add legend
            st.caption("Score ranges: Green (≥7.5) = Excellent | Yellow (5.5-7.4) = Good | Red (5.0-5.4) = Needs Improvement")
            
        except ImportError:
            # Fallback if pandas not available
            st.info("Chart visualization requires pandas. Install with: `pip install pandas`")
            # Show simple text-based ranking instead
            st.markdown("**Top Candidates by Score:**")
            for i, candidate in enumerate(sorted_for_viz[:10], 1):
                score_color = get_score_color(candidate.fit_score)
                st.markdown(f"{i}. **{candidate.name}** - <span style='color: {score_color}; font-weight: bold;'>{candidate.fit_score:.2f}/10</span>", unsafe_allow_html=True)
        except Exception as e:
            logger.error(f"Error creating visualization: {e}", exc_info=True)
            st.warning("Could not generate visualization. Showing rankings in list format instead.")
            st.markdown("**Top Candidates by Score:**")
            for i, candidate in enumerate(sorted_for_viz[:10], 1):
                score_color = get_score_color(candidate.fit_score)
                st.markdown(f"{i}. **{candidate.name}** - <span style='color: {score_color}; font-weight: bold;'>{candidate.fit_score:.2f}/10</span>", unsafe_allow_html=True)
        
        st.markdown("---")

    # Sorting and filtering controls
    st.markdown("### Candidate Rankings")

    col1, col2, col3 = st.columns([2, 2, 1])

    with col1:
        sort_option = st.selectbox(
            "Sort by",
            ["Score (High to Low)", "Score (Low to High)", "Name (A-Z)", "Name (Z-A)"],
            key="sort_option"
        )

    with col2:
        min_score = st.slider(
            "Minimum Score Filter",
            min_value=5.0,
            max_value=10.0,
            value=5.0,
            step=0.5,
            key="min_score_filter"
        )

    with col3:
        show_details = st.checkbox("Expand All", value=False, key="expand_all")

    # Apply sorting (only viable candidates)
    sorted_candidates = list(viable_candidates)
    if sort_option == "Score (High to Low)":
        sorted_candidates.sort(key=lambda x: x.fit_score, reverse=True)
    elif sort_option == "Score (Low to High)":
        sorted_candidates.sort(key=lambda x: x.fit_score)
    elif sort_option == "Name (A-Z)":
        sorted_candidates.sort(key=lambda x: x.name.lower())
    elif sort_option == "Name (Z-A)":
        sorted_candidates.sort(key=lambda x: x.name.lower(), reverse=True)

    # Apply filtering
    filtered_candidates = [c for c in sorted_candidates if c.fit_score >= min_score]

    if not sorted_candidates:
        st.info("No viable candidates (score >= 5.0).")
        return

    if len(filtered_candidates) < len(sorted_candidates):
        st.info(f"Showing {len(filtered_candidates)} of {len(sorted_candidates)} viable candidates (filtered by score >= {min_score})")

    # Display candidates with enhanced cards
    for i, candidate in enumerate(filtered_candidates, 1):
        score_color = get_score_color(candidate.fit_score)

        # Create score badge HTML
        score_badge = f"""
        <span style="
            background-color: {score_color};
            color: white;
            padding: 0.25rem 0.75rem;
            border-radius: 1rem;
            font-weight: bold;
            font-size: 1rem;
        ">{candidate.fit_score:.2f}/10</span>
        """

        # Display score in expander header
        score_display = f"{candidate.fit_score:.2f}/10"
        with st.expander(f"{i}. {candidate.name} - {score_display}", expanded=show_details):
            # Contact info - ALWAYS displayed (matching report format)
            st.markdown(f"**Email:** {candidate.email if candidate.email else 'Not provided'} • **Phone:** {candidate.phone if candidate.phone else 'Not provided'}")
            st.markdown("---")
            
            # Key Qualifications (by importance) - matching report format
            st.markdown("**Key Qualifications (by importance):**")
            
            # Get scoring weights to prioritize information
            from industry_templates import get_default_weights
            from models import JobDetails
            job_details_obj = results.get('job_details_obj')  # Try to get JobDetails object
            if not job_details_obj:
                # Fallback: create a minimal JobDetails for weights
                job_details_obj = JobDetails(
                    job_title=results['job_details']['title'],
                    certifications=[],
                    location=results['job_details']['location'],
                    full_description=""
                )
            weights = job_details_obj.scoring_profile if hasattr(job_details_obj, 'scoring_profile') and job_details_obj.scoring_profile else get_default_weights()
            
            # Get required/preferred skills from JobDetails object
            required_skills = getattr(job_details_obj, 'required_skills', []) if hasattr(job_details_obj, 'required_skills') else []
            preferred_skills = getattr(job_details_obj, 'preferred_skills', []) if hasattr(job_details_obj, 'preferred_skills') else []
            
            qual_items = []
            
            # 1. Experience Level (HIGHEST PRIORITY)
            exp_years = candidate.experience_match.get('years', 0)
            if exp_years > 0:
                qual_items.append((weights.get('experience_level', 0.25), f"• **Experience:** {exp_years} years"))
            else:
                qual_items.append((weights.get('experience_level', 0.25), f"• **Experience:** Not specified"))
            
            # 2. Job Title Match (SECOND PRIORITY)
            titles = candidate.experience_match.get('titles', [])
            if titles:
                title_list = ', '.join(titles[:3])
                if len(titles) > 3:
                    title_list += f" (+{len(titles) - 3} more)"
                qual_items.append((weights.get('job_title_match', 0.20), f"• **Job Title Match:** {title_list}"))
            else:
                qual_items.append((weights.get('job_title_match', 0.20), f"• **Job Title Match:** No relevant titles"))
            
            # 3. Required Skills (THIRD PRIORITY)
            req_skills_list = candidate.skills_match.get('candidate_skills', [])
            matched_skills = [s for s in req_skills_list if any(req.lower() in s.lower() or s.lower() in req.lower() for req in required_skills)] if required_skills else req_skills_list[:5]
            if matched_skills:
                skills_display = ', '.join(matched_skills[:5])
                if len(matched_skills) > 5:
                    skills_display += f" (+{len(matched_skills) - 5} more)"
                qual_items.append((weights.get('required_skills', 0.18), f"• **Required Skills:** {skills_display}"))
            else:
                qual_items.append((weights.get('required_skills', 0.18), f"• **Required Skills:** Limited match"))
            
            # 4. Transferrable Skills (FOURTH PRIORITY)
            transferrable_skills_list = candidate.transferrable_skills_match.get('transferrable_skills', [])
            if transferrable_skills_list:
                skills_display = ', '.join(transferrable_skills_list[:5])
                if len(transferrable_skills_list) > 5:
                    skills_display += f" (+{len(transferrable_skills_list) - 5} more)"
                qual_items.append((weights.get('transferrable_skills', 0.15), f"• **Transferrable Skills:** {skills_display}"))
            else:
                qual_items.append((weights.get('transferrable_skills', 0.15), f"• **Transferrable Skills:** Limited transferrable skills"))
            
            # 5. Location (FIFTH PRIORITY)
            if candidate.location_match:
                qual_items.append((weights.get('location', 0.10), f"• **Location:** ✓ Matches job location"))
            else:
                qual_items.append((weights.get('location', 0.10), f"• **Location:** ✗ Does not match job location"))
            
            # 6. Preferred Skills (SIXTH PRIORITY)
            pref_skills_list = candidate.skills_match.get('candidate_skills', [])
            matched_pref_skills = [s for s in pref_skills_list if any(pref.lower() in s.lower() or s.lower() in pref.lower() for pref in preferred_skills)] if preferred_skills else pref_skills_list[:5]
            if matched_pref_skills:
                skills_display = ', '.join(matched_pref_skills[:5])
                if len(matched_pref_skills) > 5:
                    skills_display += f" (+{len(matched_pref_skills) - 5} more)"
                qual_items.append((weights.get('preferred_skills', 0.07), f"• **Preferred Skills:** {skills_display}"))
            else:
                qual_items.append((weights.get('preferred_skills', 0.07), f"• **Preferred Skills:** Limited match"))
            
            # 7. Certifications/Education (LOWEST PRIORITY)
            if candidate.certifications:
                MAX_CERTS_DISPLAY = 3
                if len(candidate.certifications) > MAX_CERTS_DISPLAY:
                    cert_list = ', '.join(candidate.certifications[:MAX_CERTS_DISPLAY])
                    qual_items.append((weights.get('certifications_education', 0.05), f"• **Certifications/Education:** {cert_list} (+{len(candidate.certifications) - MAX_CERTS_DISPLAY} more)"))
                else:
                    qual_items.append((weights.get('certifications_education', 0.05), f"• **Certifications/Education:** {', '.join(candidate.certifications)}"))
            else:
                qual_items.append((weights.get('certifications_education', 0.05), f"• **Certifications/Education:** None listed"))
            
            # Sort by weight (highest first) and display
            qual_items.sort(key=lambda x: x[0], reverse=True)
            for _, bullet in qual_items:
                st.markdown(bullet)
            
            st.markdown("---")
            
            # Summary - concise 1-2 sentence fit assessment (matching report format)
            st.markdown("**Summary:**")
            from pdf_generator import PDFGenerator
            pdf_gen = PDFGenerator()
            summary_text = pdf_gen._generate_concise_summary(candidate, job_details_obj)
            st.markdown(summary_text)
            
            st.markdown("---")
            
            # AI Assessment (full rationale)
            st.markdown("**AI Assessment:**")
            st.markdown(f"<div style='background-color: var(--bg-secondary); padding: var(--spacing-sm); border-radius: 0.5rem; font-size: var(--font-size-sm);'>{candidate.rationale}</div>", unsafe_allow_html=True)

    # Show auto-save success notification if resumes were automatically saved
    if st.session_state.get('resumes_auto_saved'):
        saved_count = st.session_state.pop('resumes_auto_saved')
        st.success(f"**Success:** {saved_count} resume(s) automatically saved to your database and linked to this analysis.")
        
        st.markdown("---")
    
    # Download section
    st.markdown('<div class="section-header">Download Reports</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    # Build in-memory downloads using base64 to avoid Streamlit media cache misses
    csv_data = generate_csv_export(
        candidate_scores,
        job_details['title'],
        job_details['location']
    )
    pdf_link = _make_download_link(pdf_data, os.path.basename(pdf_path), "application/pdf")
    csv_link = _make_download_link(csv_data.encode("utf-8"), f"candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", "text/csv")

    with col1:
        if pdf_link:
            st.markdown(pdf_link, unsafe_allow_html=True)
        else:
            st.error("PDF data unavailable for download.")

    with col2:
        if csv_link:
            st.markdown(csv_link, unsafe_allow_html=True)
        else:
            st.error("CSV data unavailable for download.")

    # Only show PDF path info if PDF exists
    if pdf_path and os.path.exists(pdf_path):
        st.info(f"PDF report saved locally at: {pdf_path}")
    elif pdf_path:
        st.warning(f"PDF report file not found at: {pdf_path}. The file may have been moved or deleted.")


def main():
    # Initialize session state
    init_session_state()

    # Check for OpenAI API key at startup
    if not OpenAIConfig.is_configured():
        st.error("""
        **OpenAI API Key Not Configured**

        Please set the OPENAI_API_KEY environment variable to use this application.

        You can set it by:
        1. Adding it to your .env file: `OPENAI_API_KEY=your-key-here`
        2. Or exporting it: `export OPENAI_API_KEY=your-key-here`

        See OPENAI_API_SETUP.md for detailed instructions.
        """)
        st.stop()

    st.set_page_config(
        page_title="Recruitment Candidate Ranker",
        page_icon=None,
        layout="wide"
    )

    # Initialize theme system (before auth gate so theme toggle works)
    current_theme = get_initial_theme()
    
    # Inject theme CSS
    inject_theme_css()
    
    # Inject theme initialization script
    theme_script = render_theme_script(current_theme)
    st.markdown(theme_script, unsafe_allow_html=True)
    
    # Set data-theme attribute on html element
    st.markdown(f'<script>document.documentElement.setAttribute("data-theme", "{current_theme}");</script>', unsafe_allow_html=True)

    # Require authentication (now theme system is initialized)
    _auth_gate()
    
    # Custom CSS with logo branding colors and enhanced styling
    st.markdown("""
        <style>
        /* Logo branding colors and design system - now using design tokens */
        :root {
            /* ResponsAble Brand Colors */
            --brand-blue: #215096;  /* Deep blue from ResponsAble logo */
            --brand-green: #38A84F;  /* Bright green from ResponsAble logo */
            --brand-black: #000000;  /* Black background */
            --brand-dark-gray: #4A4A4A;  /* Dark gray for tagline */
            
            /* ResponsAble Typography */
            --font-serif: 'Times New Roman', 'Times', serif;
            --score-excellent: #27ae60;
            --score-good: #f39c12;
            --score-poor: #e74c3c;
            
            /* Spacing scale */
            --spacing-xs: 0.5rem;
            --spacing-sm: 1rem;
            --spacing-md: 1.5rem;
            --spacing-lg: 2rem;
            --spacing-xl: 3rem;
            
            /* Typography scale */
            --font-size-xs: 0.75rem;
            --font-size-sm: 0.875rem;
            --font-size-base: 1rem;
            --font-size-lg: 1.125rem;
            --font-size-xl: 1.25rem;
            --font-size-2xl: 1.5rem;
            --font-size-3xl: 1.875rem;
            
            /* Semantic colors - ResponsAble theme */
            --text-primary: #215096;  /* Blue for primary text */
            --text-secondary: #4A4A4A;
            --text-muted: #6c757d;
            --bg-primary: #ffffff;
            --bg-secondary: #f8f9fa;
            --border-color: #dee2e6;
            --border-color-hover: var(--brand-green);
            
            /* Status colors */
            --success-bg: #E8F5E9;
            --success-border: var(--brand-green);
            --info-bg: #E3F2FD;
            --info-border: var(--brand-blue);
            --warning-bg: #FFF3CD;
            --warning-border: #ffc107;
            --error-bg: #F8D7DA;
            --error-border: #DC3545;
        }
        
        /* Legacy variable mappings for backward compatibility */
        --text-primary: var(--color-text-primary);
        --text-secondary: var(--color-text-secondary);
        --text-muted: var(--color-text-muted);
        --bg-primary: var(--color-surface);
        --bg-secondary: var(--color-surface-elevated);
        --border-color: var(--color-border);
        --border-color-hover: var(--color-border-hover);
        
        /* Status colors using new tokens (will be overridden by theme) */
        --success-bg: var(--color-success);
        --success-border: var(--color-success);
        --info-bg: var(--color-info);
        --info-border: var(--color-info);
        --warning-bg: var(--color-warning);
        --warning-border: var(--color-warning);
        --error-bg: var(--color-danger);
        --error-border: var(--color-danger);

        .header-container {
            display: flex;
            align-items: center;
            gap: 1rem;
            margin-bottom: 1.5rem;
            padding-bottom: 1rem;
            border-bottom: 2px solid var(--brand-blue);  /* ResponsAble blue accent */
            font-family: var(--font-serif);
        }

        .logo-container {
            max-width: 500px;
            margin: 0 auto 1rem auto;
            text-align: center;
            display: flex;
            justify-content: center;
            align-items: center;
            width: 100%;
        }
        
        .logo-container img {
            display: block;
            margin: 0 auto;
        }

        .header-text {
            flex-grow: 1;
        }

        .main-header {
            font-size: 2.5rem;
            color: var(--brand-blue);  /* ResponsAble blue */
            font-weight: bold;
            margin-bottom: 0.3rem;
            font-family: var(--font-serif);
        }

        .main-header .blue-text {
            color: var(--brand-green);  /* ResponsAble green */
        }

        .sub-header {
            font-size: 1.2rem;
            color: var(--brand-dark-gray);
            margin-bottom: 0;
        }

        .section-header {
            font-size: var(--font-size-2xl);
            color: var(--brand-blue);  /* ResponsAble blue for headers */
            font-weight: 600;
            line-height: 1.3;
            margin-top: var(--spacing-lg);
            margin-bottom: var(--spacing-md);
            border-bottom: 2px solid var(--brand-blue);  /* ResponsAble blue accent */
            padding-bottom: var(--spacing-xs);
            font-family: var(--font-serif);
        }
        
        /* Typography improvements */
        .stMarkdown p {
            font-size: var(--font-size-base);
            line-height: 1.6;
        }
        
        .stCaption {
            font-size: var(--font-size-sm);
            color: var(--text-secondary);
        }
        
        /* Form field spacing */
        .stTextInput, .stTextArea, .stSelectbox {
            margin-bottom: var(--spacing-sm);
        }
        
        /* Card spacing */
        .stContainer {
            padding: var(--spacing-md);
            margin-bottom: var(--spacing-md);
        }
        
        /* Form section grouping */
        .form-section {
            background: var(--bg-secondary);
            padding: var(--spacing-md);
            border-radius: 0.5rem;
            margin-bottom: var(--spacing-md);
            border-left: 3px solid var(--brand-blue);
        }
        
        /* Required field indicator */
        .required-field::after {
            content: " *";
            color: var(--error-border);
            font-weight: bold;
        }
        
        /* Enhanced validation messages */
        .stError {
            display: flex;
            align-items: center;
            gap: var(--spacing-xs);
            padding: var(--spacing-sm);
            border-radius: 0.25rem;
        }

        /* Style the Process Candidates button with ResponsAble blue */
        div[data-testid="stButton"] > button[kind="primary"],
        button[kind="primary"] {
            background-color: var(--color-primary) !important;
            background: var(--color-primary) !important;
            color: var(--color-text-inverse) !important;
            border: none !important;
            font-weight: var(--font-weight-semibold) !important;
            font-size: var(--font-size-lg) !important;
            padding: var(--spacing-md) var(--spacing-xl) !important;
            border-radius: var(--radius-md) !important;
            transition: all 0.2s ease !important;
        }

        div[data-testid="stButton"] > button[kind="primary"]:hover,
        button[kind="primary"]:hover {
            background-color: var(--color-primary) !important;
            background: var(--color-primary) !important;
            opacity: 0.9;
            box-shadow: var(--shadow-md) !important;
            transform: translateY(-1px) !important;
        }
        
        div[data-testid="stButton"] > button[kind="primary"]:focus-visible,
        button[kind="primary"]:focus-visible {
            outline: 2px solid var(--color-focus-ring) !important;
            outline-offset: 2px !important;
        }

        /* Style download button with ResponsAble blue */
        div[data-testid="stDownloadButton"] > button,
        .stDownloadButton > button {
            background-color: var(--brand-blue) !important;  /* ResponsAble blue */
            background: var(--brand-blue) !important;  /* ResponsAble blue */
            color: white !important;
            border: none !important;
            font-weight: bold !important;
        }

        div[data-testid="stDownloadButton"] > button:hover,
        .stDownloadButton > button:hover {
            background-color: var(--info-border) !important;
            background: var(--info-border) !important;
        }

        /* Additional Streamlit button styling */
        button[data-baseweb="button"][kind="primary"] {
            background-color: var(--brand-blue) !important;  /* ResponsAble blue */
            background: var(--brand-blue) !important;  /* ResponsAble blue */
        }

        /* Update success messages to use brand green */
        .stSuccess {
            background-color: var(--success-bg) !important;
            border-left: 4px solid var(--success-border) !important;
            color: var(--text-primary) !important;
        }

        /* Update info boxes to use brand blue */
        .stInfo {
            background-color: var(--info-bg) !important;
            border-left: 4px solid var(--info-border) !important;
            color: var(--text-primary) !important;
        }

        /* Update error boxes */
        .stError {
            background-color: var(--error-bg) !important;
            border-left: 4px solid var(--error-border) !important;
            color: var(--text-primary) !important;
        }
        
        /* Warning boxes */
        .stWarning {
            background-color: var(--warning-bg) !important;
            border-left: 4px solid var(--warning-border) !important;
            color: var(--text-primary) !important;
        }

        /* Style file uploaders with brand colors */
        .stFileUploader > div > div {
            border: 2px dashed var(--brand-blue) !important;
        }

        .stFileUploader > div > div:hover {
            border-color: var(--brand-green) !important;
        }

        /* Style text inputs with subtle brand colors */
        .stTextInput > div > div > input:focus {
            border-color: var(--brand-green) !important;
            box-shadow: 0 0 0 2px rgba(0, 166, 81, 0.2) !important;
        }

        /* Style selectboxes */
        .stSelectbox > div > div > select:focus {
            border-color: var(--brand-green) !important;
        }

        /* Style radio buttons */
        .stRadio > div > label {
            color: var(--brand-dark-gray) !important;
        }
        
        @media (prefers-color-scheme: dark) {
            .stRadio > div > label {
                color: var(--text-primary) !important;
            }
        }

        /* Style expanders */
        .streamlit-expanderHeader {
            color: var(--brand-blue) !important;
            font-weight: 600 !important;
        }

        /* Style metrics */
        [data-testid="stMetricValue"] {
            color: var(--brand-blue) !important;
        }

        /* Progress bar styling */
        .stProgress > div > div > div {
            background-color: var(--brand-green) !important;
        }

        /* Overall page styling */
        .main .block-container {
            padding-top: var(--spacing-lg);
            padding-bottom: var(--spacing-lg);
            background-color: var(--bg-primary);
            color: var(--text-primary);
        }
        
        /* Dark mode adjustments for Streamlit components */
        @media (prefers-color-scheme: dark) {
            /* Ensure Streamlit's default styles work with dark mode */
            .stMarkdown {
                color: var(--text-primary);
            }
            
            .stMarkdown p, .stMarkdown li, .stMarkdown ul, .stMarkdown ol {
                color: var(--text-primary);
            }
            
            /* Input fields */
            .stTextInput > div > div > input,
            .stTextArea > div > div > textarea,
            .stSelectbox > div > div > select {
                background-color: var(--bg-secondary);
                color: var(--text-primary);
                border-color: var(--border-color);
            }
            
            /* Tables and dataframes */
            .stDataFrame {
                background-color: var(--bg-secondary);
                color: var(--text-primary);
            }
            
            /* Expanders */
            .streamlit-expanderHeader {
                color: var(--text-primary);
            }
            
            .streamlit-expanderContent {
                background-color: var(--bg-secondary);
                color: var(--text-primary);
            }
            
            /* Metrics */
            [data-testid="stMetricLabel"] {
                color: var(--text-secondary);
            }
            
            [data-testid="stMetricValue"] {
                color: var(--brand-blue);
            }
            
            /* Borders and dividers */
            .section-header {
                border-bottom-color: var(--brand-green);
            }
            
            .header-container {
                border-bottom-color: var(--brand-blue);
            }
            
            /* Form sections */
            .form-section {
                background-color: var(--bg-secondary);
                border-left-color: var(--brand-blue);
            }
            
            /* Loading container */
            .loading-container {
                background-color: var(--bg-secondary);
            }
            
            .loading-progress-bar {
                background-color: rgba(255, 255, 255, 0.1);
            }
            
            /* File uploader adjustments */
            .stFileUploader > div > div {
                background-color: var(--bg-secondary);
            }
            
            /* Button text contrast */
            button[kind="primary"] {
                color: white !important;  /* Ensure white text on brand green */
            }
            
            /* Download button text */
            .stDownloadButton > button {
                color: white !important;  /* Ensure white text on brand blue */
            }
            
            /* Chart/visualization adjustments */
            .stDataFrame table {
                background-color: var(--bg-secondary);
                color: var(--text-primary);
            }
            
            /* Ensure all markdown text is readable */
            .stMarkdown h1, .stMarkdown h2, .stMarkdown h3,
            .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {
                color: var(--text-primary);
            }
            
            /* Sub-header text color */
            .sub-header {
                color: var(--text-secondary);
            }
            
            /* Text input focus shadow adjustment */
            .stTextInput > div > div > input:focus {
                box-shadow: 0 0 0 2px rgba(0, 166, 81, 0.4) !important;
            }
        }
        
        /* Improved table styling */
        .stDataFrame {
            margin-bottom: var(--spacing-md);
        }
        
        /* Better metric card spacing */
        [data-testid="stMetricContainer"] {
            padding: var(--spacing-sm);
            margin-bottom: var(--spacing-sm);
        }
        
        /* Improved button spacing */
        div[data-testid="stButton"] {
            margin-bottom: var(--spacing-sm);
        }

        /* Score badge styling */
        .score-badge {
            display: inline-block;
            padding: 0.25rem 0.75rem;
            border-radius: 1rem;
            font-weight: bold;
            color: white;
        }

        .score-excellent { background-color: var(--score-excellent); }
        .score-good { background-color: var(--score-good); }
        .score-poor { background-color: var(--score-poor); }
        
        @media (prefers-color-scheme: dark) {
            /* Score colors remain the same for visibility */
            .score-excellent { background-color: var(--score-excellent); }
            .score-good { background-color: var(--score-good); }
            .score-poor { background-color: var(--score-poor); }
        }

        /* Processing status styling */
        .processing-status {
            background-color: var(--info-bg);
            border: 1px solid var(--brand-blue);
            border-radius: 0.5rem;
            padding: var(--spacing-md);
            margin: var(--spacing-md) 0;
            color: var(--text-primary);
        }

        .processing-step {
            font-weight: bold;
            color: var(--brand-blue);
            margin-bottom: var(--spacing-xs);
        }

        .processing-message {
            color: var(--text-secondary);
            font-style: italic;
        }
        
        /* Full-page loading overlay */
        .full-page-loading-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
                            background: var(--color-surface);
                            opacity: 0.98;
                            backdrop-filter: blur(2px);
            z-index: 9999;
            display: flex;
            align-items: center;
            justify-content: center;
            flex-direction: column;
            margin: 0;
            padding: 0;
            overflow: hidden;
        }
        
        @media (prefers-color-scheme: dark) {
            .full-page-loading-overlay {
                background: rgba(26, 26, 26, 0.98);
            }
        }
        
        /* Hide scrollbars during loading */
        body:has(.full-page-loading-overlay) {
            overflow: hidden;
        }
        
        /* Loading screen component styles */
        .loading-container {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 1.5rem;
            padding: 2rem;
            text-align: center;
        }
        
        .loading-spinner {
            width: 60px;
            height: 60px;
            border: 4px solid rgba(0, 166, 81, 0.2);
            border-top-color: var(--brand-green);
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }
        
        @keyframes spin {
            to { transform: rotate(360deg); }
        }
        
        .loading-title {
            font-size: var(--font-size-2xl);
            font-weight: 600;
            color: var(--brand-blue);
            margin: 0;
        }
        
        .loading-message {
            font-size: var(--font-size-base);
            color: var(--text-secondary);
            margin: 0;
            max-width: 600px;
        }
        
        .loading-progress-bar {
            width: 100%;
            max-width: 400px;
            height: 8px;
            background-color: rgba(0, 166, 81, 0.1);
            border-radius: 4px;
            overflow: hidden;
            margin-top: 0.5rem;
        }
        
        .loading-progress-fill {
            height: 100%;
            background-color: var(--brand-green);
            border-radius: 4px;
            transition: width 0.3s ease;
        }
        
        /* Responsive adjustments */
        @media (max-width: 768px) {
            .logo-container img {
                max-width: 300px;
            }
            
            .form-columns {
                flex-direction: column;
            }
            
            .section-header {
                font-size: var(--font-size-xl);
                margin-top: var(--spacing-md);
            }
        }
        </style>
    """, unsafe_allow_html=True)

    # Header with logo - ResponsAble
    logo_path = Path(__file__).parent / "responsableLOGO-color-2048px.jpg"

    # Display logo with fixed width to maintain aspect ratio while reducing size, centered
    if logo_path.exists():
        try:
            # Read logo as bytes and encode to base64 to avoid Streamlit media file storage issues
            import base64
            with open(logo_path, "rb") as f:
                logo_bytes = f.read()
                logo_base64 = base64.b64encode(logo_bytes).decode()
            
            # Use columns to center the logo
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                # Embed logo directly in HTML to avoid media file storage issues
                st.markdown(
                    f'<div style="text-align: center;"><img src="data:image/jpeg;base64,{logo_base64}" style="max-width: 400px; height: auto;" /></div>',
                    unsafe_allow_html=True
                )
        except Exception as e:
            logger.warning(f"Failed to load logo: {e}")
            # Fallback: try using st.image with error handling
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                try:
                    st.image(str(logo_path), width=400)
                except Exception:
                    pass  # Silently fail if logo can't be displayed
        st.markdown("""
            <div style="text-align: center; margin-top: 0.5rem; margin-bottom: 1rem;">
                <div style="font-size: var(--font-size-lg); color: var(--brand-blue); font-weight: 500; font-family: var(--font-serif);">
                    Safety Staffing on Demand
                </div>
            </div>
        """, unsafe_allow_html=True)
    else:
        # Fallback text logo
        col1, col2 = st.columns([1, 4])

        with col1:
            st.markdown("""
                <div style="font-size: var(--font-size-2xl); font-weight: bold; color: var(--brand-blue); font-family: 'Times New Roman', serif;">
                    RA
                </div>
            """, unsafe_allow_html=True)

        with col2:
            st.markdown("""
                <div class="header-text">
                    <div class="main-header">
                        <span style="color: var(--brand-blue); font-family: 'Times New Roman', serif; font-weight: bold;">RESPONS</span><span style="color: var(--brand-green); font-family: 'Times New Roman', serif; font-weight: bold;">ABLE</span>
                    </div>
                    <div class="sub-header">Safety Staffing on Demand</div>
                </div>
            """, unsafe_allow_html=True)

    # Main page tabs: New Analysis and Previous Analyses
    user_id = st.session_state.get("user_id")
    if user_id:
        tab_new, tab_history, tab_database, tab_analytics = st.tabs(["New Analysis", "Previous Analyses", "Resume Database", "Analytics"])
        
        with tab_new:
            # Check if processing is in progress - show full-page loading screen
            if st.session_state.get('processing', False):
                # Full-page loading screen - read progress from session state
                loading_placeholder = st.empty()
                with loading_placeholder.container():
                    # Get current progress from session state
                    progress_info = st.session_state.get('loading_progress', {
                        'step': 'analyzing',
                        'progress': 0.0,
                        'current': 0,
                        'total': 0
                    })
                    
                    # Use full viewport styling with all necessary CSS
                    st.markdown("""
                        <style>
                        .full-page-loading-overlay {
                            position: fixed;
                            top: 0;
                            left: 0;
                            width: 100vw;
                            height: 100vh;
                            background: var(--color-surface);
                            opacity: 0.98;
                            backdrop-filter: blur(2px);
                            z-index: 9999;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            flex-direction: column;
                            margin: 0;
                            padding: 0;
                            overflow: hidden;
                        }
                        .loading-container {
                            display: flex;
                            flex-direction: column;
                            align-items: center;
                            justify-content: center;
                            gap: 1.5rem;
                            padding: 2rem;
                            text-align: center;
                        }
                        .loading-spinner {
                            width: 60px;
                            height: 60px;
                            border: 4px solid rgba(0, 166, 81, 0.2);
                            border-top-color: var(--color-primary);
                            border-radius: 50%;
                            animation: spin 1s linear infinite;
                        }
                        @keyframes spin {
                            to { transform: rotate(360deg); }
                        }
                        .loading-title {
                            font-size: 1.5rem;
                            font-weight: 600;
                            color: var(--color-text-primary);
                            margin: 0;
                        }
                        .loading-message {
                            font-size: 1rem;
                            color: var(--color-text-muted);
                            margin: 0;
                            max-width: 600px;
                        }
                        .loading-progress-bar {
                            width: 100%;
                            max-width: 400px;
                            height: 8px;
                            background-color: rgba(0, 166, 81, 0.1);
                            border-radius: 4px;
                            overflow: hidden;
                            margin-top: 0.5rem;
                        }
                        .loading-progress-fill {
                            height: 100%;
                            background-color: var(--color-primary);
                            border-radius: 4px;
                            transition: width 0.3s ease;
                        }
                        </style>
                    """, unsafe_allow_html=True)
                    st.markdown('<div class="full-page-loading-overlay">', unsafe_allow_html=True)
                    from loading_components import show_full_workflow_loading
                    show_full_workflow_loading(
                        progress_info['step'],
                        progress_info['progress'],
                        progress_info['current'],
                        progress_info['total']
                    )
                    st.markdown('</div>', unsafe_allow_html=True)
            # Check if we should show results page (after processing completes)
            elif st.session_state.get('show_results_page', False) and st.session_state.results is not None:
                display_results(st.session_state.results)
                
                # Footer
                st.markdown("---")
                st.markdown(
                    '<div style="text-align: center; color: var(--text-muted); padding: var(--spacing-sm);">ResponsAble | Safety Staffing on Demand</div>',
                    unsafe_allow_html=True
                )
            # If we have stored results from viewing previous analysis, show them
            elif st.session_state.results is not None and not st.session_state.get("viewing_previous_analysis", False):
                display_results(st.session_state.results)
                
                # Footer
                st.markdown("---")
                st.markdown(
                    '<div style="text-align: center; color: var(--text-muted); padding: var(--spacing-sm);">ResponsAble | Safety Staffing on Demand</div>',
                    unsafe_allow_html=True
                )
            else:
                # Introduction
                with st.expander("How It Works", expanded=False):
                    st.markdown("""
                    This application analyzes job requirements and candidate resumes to provide intelligent rankings:

                    **Input Methods**:
                    - **Job Sources**: Upload job description file (AI auto-extracts everything) or enter manually
                    - **Candidate Sources**: Upload resume files or select from your saved database

                    **Processing Steps**:
                    1. **Job Analysis**: Extracts requirements from job description (automatically if uploaded file)
                    2. **Research**: Identifies equivalent skills and job titles
                    3. **Resume Parsing**: Extracts structured data from resumes
                    4. **Scoring**: Evaluates candidates with AI-powered analysis and chain-of-thought reasoning
                    5. **Ranking**: Selects top 4-10 candidates
                    6. **Report**: Generates professional PDF with visualizations

                    **Default Scoring Criteria** (Universal Standard - in order of priority):
                    1. **Experience Level**: 25% - Candidate's years and depth of relevant experience
                    2. **Job Title Match**: 20% - Relevance of candidate's previous job titles to the role
                    3. **Required Skills**: 18% - Skills explicitly listed as required for the position
                    4. **Transferrable Skills**: 15% - AI-identified skills from candidate's experience that are relevant but not explicitly listed (cross-industry, semantically similar, or relevant unlisted skills)
                    5. **Location**: 10% - Geographic proximity to job location
                    6. **Preferred Skills**: 7% - Skills listed as preferred (nice-to-have)
                    7. **Certifications/Education**: 5% - Certifications, licenses, degrees, and education level (combined evaluation)

                    **Customizable Scoring**:
                    - **Industry Templates**: Choose from pre-configured templates (General, Healthcare, Technology, Construction, Finance, Sales) that adjust weights based on industry priorities
                    - **Custom Weights**: Manually adjust the importance of each criterion to match your specific hiring needs
                    - **Dealbreakers**: Set criteria that automatically disqualify candidates
                    - **Bias Reduction**: Enable blind screening to reduce unconscious bias

                    **AI-Powered Features**:
                    - **Transferrable Skills Analysis**: The AI identifies skills from a candidate's experience that are relevant to the role but not explicitly listed in job requirements
                    - **Semantic Understanding**: Recognizes skills with different wording but same meaning
                    - **Cross-Industry Skills**: Identifies transferrable abilities from related fields
                    """)

            # Job Source Selection
            st.markdown('<div class="section-header">Job Source</div>', unsafe_allow_html=True)
            
            input_mode = st.radio(
                "Choose how to provide job details:",
                ["Upload Job Description File (AI Extracts Everything)", "Manual Entry"],
                horizontal=True,
                help="File upload uses AI to automatically extract all job details.",
                key="input_mode_radio"
            )
            st.caption("**Note:** Job details are required. Candidate resumes are optional - you can add them later or upload them now.")
            
            # Clear edited values when switching modes
            if 'last_input_mode' in st.session_state and st.session_state['last_input_mode'] != input_mode:
                st.session_state.pop('edited_job_title', None)
                st.session_state.pop('edited_location', None)
                st.session_state.pop('edited_certifications', None)
                st.session_state.pop('edited_job_description', None)
                st.session_state.pop('edited_required_skills', None)
                st.session_state.pop('edited_preferred_skills', None)
                st.session_state.pop('edited_experience_level', None)
                st.session_state.pop('edit_extracted_info', None)
                st.session_state.pop('extracted_job_data', None)
                st.session_state.pop('last_uploaded_file_hash', None)  # Clear hash when input mode changes
                st.session_state.pop('current_required_skills', None)
                st.session_state.pop('current_preferred_skills', None)
            st.session_state['last_input_mode'] = input_mode

            job_title = ""
            location = ""
            certifications = []
            job_description = ""

            # MODE 1: File Upload (AI Auto-Extraction)
            if input_mode == "Upload Job Description File (AI Extracts Everything)":
                st.markdown('<div class="section-header">Upload Job Description</div>', unsafe_allow_html=True)

                job_desc_file = st.file_uploader(
                    "Upload Job Description File",
                    type=["txt", "pdf", "docx"],
                    help="Upload job description - AI will automatically extract job title, location, certifications, and all requirements"
                )

                if job_desc_file:
                    # Validate file size (200MB limit)
                    max_size_mb = 200
                    max_size_bytes = max_size_mb * 1024 * 1024
                    if job_desc_file.size > max_size_bytes:
                        st.error(f"**Error - File too large:** '{job_desc_file.name}' is {job_desc_file.size / (1024*1024):.1f}MB. Maximum file size is {max_size_mb}MB. Please upload a smaller file.")
                        job_desc_file = None
                    
                    if job_desc_file:
                        import hashlib
                        
                        # Calculate file hash for more reliable tracking (name + size + content hash)
                        # Use getvalue() instead of getbuffer() for consistent hashing across reruns
                        file_bytes = job_desc_file.getvalue()
                        file_hash = hashlib.md5(file_bytes).hexdigest()
                        file_size = len(file_bytes)
                        file_identifier = f"{job_desc_file.name}_{file_size}_{file_hash[:8]}"
                        
                        # Check if this exact file has already been parsed - skip parsing if already done
                        file_already_parsed = (
                            'extracted_job_data' in st.session_state and 
                            st.session_state.get('last_uploaded_file_hash') == file_hash
                        )
                        
                        # Clear previous edited values when new file is uploaded
                        if not file_already_parsed:
                            # Clear edited values to start fresh
                            st.session_state.pop('edited_job_title', None)
                            st.session_state.pop('edited_location', None)
                            st.session_state.pop('edited_certifications', None)
                            st.session_state.pop('edited_job_description', None)
                            st.session_state.pop('edited_required_skills', None)
                            st.session_state.pop('edited_preferred_skills', None)
                            st.session_state.pop('edited_experience_level', None)
                            st.session_state.pop('edit_extracted_info', None)
                            st.session_state['last_uploaded_file'] = job_desc_file.name
                            # Don't store hash here - store it AFTER successful parsing
                        
                        # Only parse if file hasn't been parsed yet
                        if not file_already_parsed:
                            # Show loading screen with progress bar
                            loading_placeholder = st.empty()
                            try:
                                # Step 1: Reading file and preparing content (0-20%)
                                with loading_placeholder.container():
                                    from loading_components import show_job_analysis_loading
                                    show_job_analysis_loading(progress=0.1)
                                
                                # Save to temp file for parser
                                temp_dir = tempfile.mkdtemp()
                                temp_path = os.path.join(temp_dir, job_desc_file.name)

                                try:
                                    with open(temp_path, 'wb') as f:
                                        f.write(file_bytes)  # Use file_bytes already calculated for hash
                                    
                                    # Step 2: File read complete (20%)
                                    with loading_placeholder.container():
                                        show_job_analysis_loading(progress=0.2)

                                    # Step 3: AI extraction (20-60%)
                                    with loading_placeholder.container():
                                        show_job_analysis_loading(progress=0.3)
                                    
                                    # Parse the job description file using AI parser
                                    parser = AIJobParser()  # Uses AI if API key available
                                    
                                    # Update progress during AI call (30-50%)
                                    with loading_placeholder.container():
                                        show_job_analysis_loading(progress=0.5)
                                    
                                    job_data = parser.parse(temp_path)
                                    
                                    # Step 4: Parsing AI response (60-80%)
                                    with loading_placeholder.container():
                                        show_job_analysis_loading(progress=0.7)
                                finally:
                                    # Cleanup temporary files
                                    import shutil
                                    try:
                                        shutil.rmtree(temp_dir)
                                    except Exception as e:
                                        logger.warning(f"Failed to cleanup temp directory {temp_dir}: {e}", exc_info=True)
                                        # Continue - cleanup errors are non-critical

                                # Step 5: Storing results and initializing form fields (80-100%)
                                with loading_placeholder.container():
                                    show_job_analysis_loading(progress=0.9)

                                # Store extracted data in session state for editing
                                st.session_state['extracted_job_data'] = job_data
                                # Store hash AFTER successful parsing to ensure it persists
                                st.session_state['last_uploaded_file_hash'] = file_hash  # Use hash calculated at start of function
                                st.session_state['last_uploaded_file'] = job_desc_file.name
                                
                                # Initialize edited values if not already set
                                if 'edited_job_title' not in st.session_state:
                                    st.session_state['edited_job_title'] = job_data.get('job_title', '')
                                if 'edited_location' not in st.session_state:
                                    st.session_state['edited_location'] = job_data.get('location', '')
                                if 'edited_certifications' not in st.session_state:
                                    st.session_state['edited_certifications'] = job_data.get('certifications', [])
                                if 'edited_job_description' not in st.session_state:
                                    st.session_state['edited_job_description'] = job_data.get('full_description', '')
                                if 'edited_required_skills' not in st.session_state:
                                    st.session_state['edited_required_skills'] = job_data.get('required_skills', [])
                                if 'edited_preferred_skills' not in st.session_state:
                                    st.session_state['edited_preferred_skills'] = job_data.get('preferred_skills', [])
                                if 'edited_experience_level' not in st.session_state:
                                    st.session_state['edited_experience_level'] = job_data.get('experience_level', '')
                                
                                # Hide loading screen
                                loading_placeholder.empty()
                                
                                # Show success with extracted info
                                st.success(f"Successfully processed: {job_desc_file.name}")
                            except Exception as e:
                                # Hide loading screen on error
                                loading_placeholder.empty()
                                logger.error(f"Error parsing job description: {e}", exc_info=True)
                                
                                # Provide more helpful error messages
                                error_msg = str(e)
                                if "400" in error_msg or "Bad Request" in error_msg:
                                    st.error(f"**Error - File Processing:** The file '{job_desc_file.name}' could not be processed. This may be due to:\n\n"
                                            "- File format issue (corrupted or unsupported format)\n"
                                            "- File is too large or contains invalid content\n"
                                            "- PDF is image-based (scanned) rather than text-based\n\n"
                                            "**Please try:**\n"
                                            "- Converting the PDF to text format\n"
                                            "- Using a different file format (DOCX or TXT)\n"
                                            "- Ensuring the file is not corrupted")
                                elif "API" in error_msg or "OpenAI" in error_msg or "key" in error_msg.lower():
                                    st.error(f"**Error - API:** {error_msg}\n\n"
                                            "Please check your OpenAI API key configuration.")
                                elif "read" in error_msg.lower() or "parse" in error_msg.lower() or "encrypted" in error_msg.lower():
                                    st.error(f"**Error - File Reading:** Could not read the file '{job_desc_file.name}'. "
                                            "The file may be corrupted, encrypted, or in an unsupported format. Please try a different file.")
                                else:
                                    st.error(f"**Error processing file:** {error_msg}\n\n"
                                            f"If this problem persists, please try uploading a different file format (TXT, DOCX, or text-based PDF).")
                                
                                # Clear the file from session state on error so user can retry
                                if 'last_uploaded_file' in st.session_state:
                                    st.session_state.pop('last_uploaded_file', None)
                                    st.session_state.pop('last_uploaded_file_hash', None)
                                    st.session_state.pop('extracted_job_data', None)
                                # Set job_data to empty dict to prevent errors downstream
                                job_data = {}
                        
                        # If file was already parsed, use cached data (only if parsing didn't happen/fail)
                        if file_already_parsed:
                            # Use cached extracted data
                            job_data = st.session_state.get('extracted_job_data', {})
                            
                            # Ensure edited values are initialized from cached data
                            if 'edited_job_title' not in st.session_state and job_data.get('job_title'):
                                st.session_state['edited_job_title'] = job_data.get('job_title', '')
                            if 'edited_location' not in st.session_state and job_data.get('location'):
                                st.session_state['edited_location'] = job_data.get('location', '')
                            if 'edited_certifications' not in st.session_state and job_data.get('certifications'):
                                st.session_state['edited_certifications'] = job_data.get('certifications', [])
                            if 'edited_job_description' not in st.session_state and job_data.get('full_description'):
                                st.session_state['edited_job_description'] = job_data.get('full_description', '')
                            if 'edited_required_skills' not in st.session_state and job_data.get('required_skills'):
                                st.session_state['edited_required_skills'] = job_data.get('required_skills', [])
                            if 'edited_preferred_skills' not in st.session_state and job_data.get('preferred_skills'):
                                st.session_state['edited_preferred_skills'] = job_data.get('preferred_skills', [])
                            if 'edited_experience_level' not in st.session_state and job_data.get('experience_level'):
                                st.session_state['edited_experience_level'] = job_data.get('experience_level', '')
                        
                        # Ensure job_data is defined (for both parsed and cached paths)
                        if 'job_data' not in locals() or not job_data:
                            job_data = st.session_state.get('extracted_job_data', {})
                        
                        # Only proceed with edit/view mode if we have valid job_data
                        if job_data:
                            # Check if we're in edit mode (for both parsed and cached data)
                            edit_mode = st.session_state.get("edit_extracted_info", False)
                            
                            if edit_mode:
                                # EDIT MODE: Show editable form fields
                                st.markdown("### Edit Extracted Information")
                                
                                # Job Title
                                edited_job_title = st.text_input(
                                    "Job Title",
                                    value=st.session_state.get('edited_job_title', job_data.get('job_title', '')),
                                    key="edit_job_title",
                                    help="Edit the job title"
                                )
                                
                                # Location
                                edited_location = st.text_input(
                                    "Location",
                                    value=st.session_state.get('edited_location', job_data.get('location', '')),
                                    key="edit_location",
                                    help="Edit the location"
                                )
                                
                                # Experience Level
                                exp_level = st.session_state.get('edited_experience_level', job_data.get('experience_level', ''))
                                exp_options = ["", "Junior", "Mid", "Senior"]
                                exp_index = exp_options.index(exp_level) if exp_level in exp_options else 0
                                edited_experience_level = st.selectbox(
                                    "Experience Level",
                                    options=exp_options,
                                    index=exp_index,
                                    key="edit_experience_level",
                                    help="Select experience level"
                                )
                                
                                # Certifications - convert to editable text format
                                current_certs = st.session_state.get('edited_certifications', job_data.get('certifications', []))
                                cert_text_lines = []
                                for cert in current_certs:
                                    if cert.get('category') == 'must-have':
                                        cert_text_lines.append(f"*{cert.get('name', '')}")
                                    else:
                                        cert_text_lines.append(cert.get('name', ''))
                                cert_text_default = "\n".join(cert_text_lines)
                                
                                st.markdown("**Certifications** (one per line, prefix with `*` for required)")
                                st.caption("Example: `*OSHA 30` (required) or `First Aid` (preferred)")
                                
                                edited_cert_text = st.text_area(
                                    "Certifications",
                                    value=cert_text_default,
                                    height=100,
                                    key="edit_certifications_text",
                                    help="Enter one certification per line. Prefix with * for must-have/required certifications.",
                                    label_visibility="collapsed"
                                )
                                
                                # Parse certifications from text
                                edited_certifications = []
                                if edited_cert_text:
                                    for line in edited_cert_text.strip().split('\n'):
                                        line = line.strip()
                                        if line:
                                            if line.startswith('*'):
                                                edited_certifications.append({
                                                    "name": line[1:].strip(),
                                                    "category": "must-have"
                                                })
                                            else:
                                                edited_certifications.append({
                                                    "name": line,
                                                    "category": "bonus"
                                                })
                                
                                if edited_certifications:
                                    st.caption(f"Parsed: {len([c for c in edited_certifications if c['category'] == 'must-have'])} required, {len([c for c in edited_certifications if c['category'] == 'bonus'])} preferred")
                                
                                # Required Skills
                                current_req_skills = st.session_state.get('edited_required_skills', job_data.get('required_skills', []))
                                edited_required_skills_text = st.text_area(
                                    "Required Skills",
                                    value=", ".join(current_req_skills) if current_req_skills else "",
                                    height=80,
                                    key="edit_required_skills",
                                    help="Comma-separated list of required skills"
                                )
                                edited_required_skills = [s.strip() for s in edited_required_skills_text.split(',') if s.strip()] if edited_required_skills_text else []
                                
                                # Preferred Skills
                                current_pref_skills = st.session_state.get('edited_preferred_skills', job_data.get('preferred_skills', []))
                                edited_preferred_skills_text = st.text_area(
                                    "Preferred Skills",
                                    value=", ".join(current_pref_skills) if current_pref_skills else "",
                                    height=80,
                                    key="edit_preferred_skills",
                                    help="Comma-separated list of preferred skills"
                                )
                                edited_preferred_skills = [s.strip() for s in edited_preferred_skills_text.split(',') if s.strip()] if edited_preferred_skills_text else []
                                
                                # Job Description
                                edited_job_description = st.text_area(
                                    "Full Job Description",
                                    value=st.session_state.get('edited_job_description', job_data.get('full_description', '')),
                                    height=300,
                                    key="edit_job_description",
                                    help="Edit the full job description"
                                )
                                
                                # Save/Cancel buttons
                                col_save, col_cancel = st.columns([1, 1])
                                with col_save:
                                    if st.button("Save Changes", type="primary", key="save_edited_info", width='stretch'):
                                        # Store edited values
                                        st.session_state['edited_job_title'] = edited_job_title
                                        st.session_state['edited_location'] = edited_location
                                        st.session_state['edited_certifications'] = edited_certifications
                                        st.session_state['edited_job_description'] = edited_job_description
                                        st.session_state['edited_required_skills'] = edited_required_skills
                                        st.session_state['edited_preferred_skills'] = edited_preferred_skills
                                        st.session_state['edited_experience_level'] = edited_experience_level
                                        st.session_state["edit_extracted_info"] = False
                                        st.success("Changes saved!")
                                        st.rerun()
                                
                                with col_cancel:
                                    if st.button("Cancel", key="cancel_edit", width='stretch'):
                                        st.session_state["edit_extracted_info"] = False
                                        st.rerun()
                            else:
                                # VIEW MODE: Show read-only display
                                st.markdown("### AI-Extracted Information")
                                
                                # Use edited values if they exist, otherwise use extracted values
                                display_job_title = st.session_state.get('edited_job_title', job_data.get('job_title', ''))
                                display_location = st.session_state.get('edited_location', job_data.get('location', ''))
                                display_certifications = st.session_state.get('edited_certifications', job_data.get('certifications', []))
                                display_job_description = st.session_state.get('edited_job_description', job_data.get('full_description', ''))
                                display_experience_level = st.session_state.get('edited_experience_level', job_data.get('experience_level', ''))

                                col1, col2 = st.columns(2)
                                with col1:
                                    st.metric("Job Title", display_job_title if display_job_title else "Not found")
                                    st.metric("Location", display_location if display_location else "Not found")
                                    if display_experience_level:
                                        st.metric("Experience Level", display_experience_level)
                                with col2:
                                    st.metric("Certifications Found", len(display_certifications))
                                    st.metric("Document Length", f"{len(display_job_description)} chars")

                                # Show certifications if found
                                if display_certifications:
                                    with st.expander("View Extracted Certifications"):
                                        for i, cert in enumerate(display_certifications, 1):
                                            category_label = "[Required]" if cert.get('category') == 'must-have' else "[Preferred]"
                                            st.write(f"{i}. {category_label} - **{cert.get('name')}**")

                                # Show preview
                                with st.expander("View Full Job Description"):
                                    st.text_area("Content", display_job_description, height=300, disabled=True, key="jd_preview")
                                
                                    # Edit button
                                    if st.button("Edit Extracted Information", key="edit_extracted_btn"):
                                        st.session_state["edit_extracted_info"] = True
                                        st.rerun()
                                
                                # Use edited values for processing (will be used later)
                                # These will be used when the process button is clicked
                                if 'edited_job_title' in st.session_state:
                                    job_title = st.session_state.get('edited_job_title', job_data.get('job_title', ''))
                                else:
                                    job_title = job_data.get('job_title', '')
                                
                                if 'edited_location' in st.session_state:
                                    location = st.session_state.get('edited_location', job_data.get('location', ''))
                                else:
                                    location = job_data.get('location', '')
                                
                                if 'edited_certifications' in st.session_state:
                                    certifications = st.session_state.get('edited_certifications', job_data.get('certifications', []))
                                else:
                                    certifications = job_data.get('certifications', [])
                                
                                if 'edited_job_description' in st.session_state:
                                    job_description = st.session_state.get('edited_job_description', job_data.get('full_description', ''))
                                else:
                                    job_description = job_data.get('full_description', '')
                                
                                # Store edited skills for later use
                                if 'edited_required_skills' in st.session_state:
                                    st.session_state['current_required_skills'] = st.session_state.get('edited_required_skills', job_data.get('required_skills', []))
                                else:
                                    st.session_state['current_required_skills'] = job_data.get('required_skills', [])
                                
                                if 'edited_preferred_skills' in st.session_state:
                                    st.session_state['current_preferred_skills'] = st.session_state.get('edited_preferred_skills', job_data.get('preferred_skills', []))
                                else:
                                    st.session_state['current_preferred_skills'] = job_data.get('preferred_skills', [])

            # MODE 2: Manual Entry
            else:
                st.markdown('<div class="section-header">Job Details</div>', unsafe_allow_html=True)
                st.caption("Fields marked with * are required.")

                col1, col2 = st.columns(2)

                with col1:
                    job_title = st.text_input(
                        "Job Title *",
                        placeholder="e.g., Safety Specialist",
                        help="Required: Enter the exact job title"
                    )

                with col2:
                    location = st.text_input(
                        "Location *",
                        placeholder="e.g., Houston, TX or Remote",
                        help="Required: Enter the job location"
                    )

                # Simplified Certifications Entry
                st.markdown(f"<div style='margin-top: var(--spacing-md); margin-bottom: var(--spacing-xs);'><strong>Certifications</strong> (one per line, prefix with <code>*</code> for required)</div>", unsafe_allow_html=True)
                st.caption("Example: `*OSHA 30` (required) or `First Aid` (preferred)")

                cert_text = st.text_area(
                    "Certifications",
                    height=100,
                    placeholder="*OSHA 30\n*BCSP Certification\nFirst Aid\nCPR",
                    help="Enter one certification per line. Prefix with * for must-have/required certifications.",
                    label_visibility="collapsed"
                )

                # Parse certifications from text
                certifications = []
                if cert_text:
                    for line in cert_text.strip().split('\n'):
                        line = line.strip()
                        if line:
                            if line.startswith('*'):
                                certifications.append({
                                    "name": line[1:].strip(),
                                    "category": "must-have"
                                })
                            else:
                                certifications.append({
                                    "name": line,
                                    "category": "bonus"
                                })

                if certifications:
                    st.caption(f"Parsed: {len([c for c in certifications if c['category'] == 'must-have'])} required, {len([c for c in certifications if c['category'] == 'bonus'])} preferred")

                # Job Description
                st.markdown('<div class="section-header">Job Description</div>', unsafe_allow_html=True)

                job_description = st.text_area(
                    "Full Job Description *",
                    height=300,
                    placeholder="""Paste the complete job description including:
- Required skills and experience
- Preferred qualifications
- Responsibilities
- Technical requirements
- Any other relevant details

The AI will analyze this to extract skills and requirements.""",
                    help="Required: Provide as much detail as possible for accurate matching"
                )

            # Universal Recruiting Tool Enhancements
            st.markdown('<div class="section-header">Scoring Configuration (Optional)</div>', unsafe_allow_html=True)
            st.caption("Customize how candidates are evaluated. Leave defaults for standard scoring.")
            
            with st.expander("⚙️ Advanced Scoring Options", expanded=False):
                from industry_templates import get_industry_templates, get_default_weights
                from scoring_profiles import validate_weights
                
                # Industry Template Selector
                templates = get_industry_templates()
                template_names = {name: template.description for name, template in templates.items()}
                
                selected_template = st.selectbox(
                    "Industry Template",
                    options=["general"] + [k for k in templates.keys() if k != "general"],
                    format_func=lambda x: f"{x.title()} - {templates[x].description}" if x in templates else x.title(),
                    help="Select a pre-configured scoring profile for your industry. Defaults to General (universal standard). You can customize weights below.",
                    index=0,  # Default to "general" (first option)
                    key="industry_template_select"
                )
                
                # Show template preview
                if selected_template in templates:
                    template = templates[selected_template]
                    st.info(f"**{template.name.title()} Template**: {template.description}")
                    
                    # Display weights as a simple visualization
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**Scoring Weights:**")
                        for key, weight in template.weights.items():
                            weight_pct = weight * 100
                            st.caption(f"• {key.replace('_', ' ').title()}: {weight_pct:.0f}%")
                    
                    # Apply template button
                    if st.button("Apply Template", key="apply_template_btn"):
                        st.session_state['selected_template'] = selected_template
                        st.session_state['custom_weights'] = template.weights.copy()
                        st.success(f"Applied {template.name.title()} template!")
                
                st.markdown("---")
                
                # Custom Scoring Weights Editor
                st.markdown("**Custom Scoring Weights** (optional - overrides template)")
                st.caption("Adjust the importance of each evaluation criterion. Weights must sum to 100%.")
                
                # Get current weights (from template or session state)
                # Default to General template weights (universal standard)
                if 'custom_weights' in st.session_state and st.session_state['custom_weights']:
                    current_weights = st.session_state['custom_weights']
                elif selected_template in templates:
                    current_weights = templates[selected_template].weights.copy()
                else:
                    # Default to General/Universal template weights (universal standard)
                    current_weights = templates['general'].weights.copy()
                
                # Weight sliders (NEW PRIORITY ORDER)
                st.caption("Weights are ordered by default priority. Adjust to customize evaluation focus.")
                
                # 1. Experience Level (HIGHEST PRIORITY)
                weight_experience = st.slider(
                    "1. Experience Level (Highest Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('experience_level', 0.25),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for candidate's years and level of experience",
                    key="weight_experience"
                )
                
                # 2. Job Title Match (SECOND PRIORITY)
                weight_job_title = st.slider(
                    "2. Job Title Match (Second Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('job_title_match', 0.20),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for relevance of candidate's previous job titles",
                    key="weight_job_title"
                )
                
                # 3. Required Skills (THIRD PRIORITY)
                weight_required_skills = st.slider(
                    "3. Required Skills (Third Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('required_skills', 0.18),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for skills explicitly listed as required for the role",
                    key="weight_required_skills"
                )
                
                # 4. Transferrable Skills (FOURTH PRIORITY - NEW)
                weight_transferrable_skills = st.slider(
                    "4. Transferrable Skills (Fourth Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('transferrable_skills', 0.15),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for AI-identified skills from candidate's experience that are relevant but not explicitly listed (cross-industry, semantically similar, or relevant unlisted skills)",
                    key="weight_transferrable_skills"
                )
                
                # 5. Location (FIFTH PRIORITY)
                weight_location = st.slider(
                    "5. Location (Fifth Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('location', 0.10),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for geographic location match",
                    key="weight_location"
                )
                
                # 6. Preferred Skills (SIXTH PRIORITY)
                weight_preferred_skills = st.slider(
                    "6. Preferred Skills (Sixth Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('preferred_skills', 0.07),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for skills listed as preferred (nice-to-have) for the role",
                    key="weight_preferred_skills"
                )
                
                # 7. Certifications/Education (LOWEST PRIORITY - COMBINED)
                weight_certifications_education = st.slider(
                    "7. Certifications/Education (Lowest Priority)",
                    min_value=0.0,
                    max_value=1.0,
                    value=current_weights.get('certifications_education', 0.05),
                    step=0.01,
                    format="%.0f%%",
                    help="Weight for certifications, licenses, degrees, and education level (combines must-have certs, bonus certs, and education)",
                    key="weight_certifications_education"
                )
                
                # Calculate total
                total_weight = (weight_experience + weight_job_title + weight_required_skills + 
                              weight_transferrable_skills + weight_location + weight_preferred_skills + 
                              weight_certifications_education)
                
                # Validation
                if abs(total_weight - 1.0) > 0.01:
                    st.error(f"⚠️ Weights must sum to 100%. Current total: {total_weight*100:.1f}%")
                    st.caption("Adjust the sliders above to fix the total.")
                else:
                    st.success(f"✓ Weights valid (Total: {total_weight*100:.0f}%)")
                    # Store custom weights (NEW STRUCTURE)
                    custom_weights = {
                        'experience_level': weight_experience,
                        'job_title_match': weight_job_title,
                        'required_skills': weight_required_skills,
                        'transferrable_skills': weight_transferrable_skills,
                        'location': weight_location,
                        'preferred_skills': weight_preferred_skills,
                        'certifications_education': weight_certifications_education
                    }
                    st.session_state['custom_weights'] = custom_weights
                
                st.markdown("---")
                
                # Dealbreakers Section
                st.markdown("**Dealbreakers** (optional)")
                st.caption("Candidates meeting these criteria will be automatically disqualified (score = 0.0)")
                
                dealbreaker_text = st.text_area(
                    "Dealbreaker Criteria",
                    height=100,
                    placeholder="One per line, e.g.:\nMissing required license\nNo relevant experience\nCriminal record",
                    help="Enter criteria that automatically disqualify candidates. One per line.",
                    key="dealbreakers_text"
                )
                
                dealbreakers = []
                if dealbreaker_text:
                    dealbreakers = [line.strip() for line in dealbreaker_text.strip().split('\n') if line.strip()]
                    if dealbreakers:
                        st.caption(f"Configured {len(dealbreakers)} dealbreaker(s)")
                
                st.markdown("---")
                
                # Bias Reduction Toggle
                bias_reduction = st.checkbox(
                    "Enable Bias Reduction (Blind Screening)",
                    value=False,
                    help="When enabled, removes names, photos, and graduation dates from evaluation to reduce unconscious bias",
                    key="bias_reduction_checkbox"
                )
                
                if bias_reduction:
                    st.info("ℹ️ Bias reduction enabled. Personal information will be excluded from scoring.")

            # Candidate Source Selection
            st.markdown('<div class="section-header">Candidate Source (Optional)</div>', unsafe_allow_html=True)
            candidate_source = st.radio(
                "Choose candidate source:",
                ["Upload Resumes", "Select from Database"],
                horizontal=True,
                key="candidate_source_radio",
                help="Choose how to provide candidate information"
            )
            
            uploaded_files = []
            selected_candidate_ids = []
            
            if candidate_source == "Select from Database":
                # Section 1: Select from Database (Optional)
                st.markdown("**Select candidates from your resume database (optional):**")
                all_candidates = search_candidates(user_id, query="", filters={})
                
                if all_candidates:
                    candidate_options = {f"{c['name']} ({c['email']})": c['id'] for c in all_candidates}
                    selected_names = st.multiselect(
                        "Choose candidates",
                        options=list(candidate_options.keys()),
                        key="selected_candidate_names",
                        help="Optional: Select one or more candidates from your database"
                    )
                    selected_candidate_ids = [candidate_options[name] for name in selected_names]
                    
                    if selected_candidate_ids:
                        st.success(f"Selected {len(selected_candidate_ids)} candidate(s) from database")
                        with st.expander("View selected candidates", expanded=False):
                            selected_profiles = get_candidates_for_analysis(user_id, selected_candidate_ids)
                            for profile in selected_profiles:
                                st.write(f"- {profile['name']} ({profile['email']})")
                else:
                    st.info("No candidates in database. Upload resumes below to add them.")
                
                # Section 2: Upload Additional Resumes (Optional)
                st.markdown("**Upload additional resume files (optional):**")
                uploaded_files = st.file_uploader(
                    "Upload Resume Files",
                    type=["pdf", "docx", "txt"],
                    accept_multiple_files=True,
                    help="Optional: Upload candidate resumes in PDF, DOCX, or TXT format. These will be automatically saved to your database after analysis.",
                    key="resume_uploader"
                )

                if uploaded_files:
                    st.success(f"Uploaded {len(uploaded_files)} resume(s)")
                    with st.expander("View uploaded files", expanded=False):
                        for file in uploaded_files:
                            size_kb = file.size / 1024
                            st.write(f"- {file.name} ({size_kb:.1f} KB)")
            else:
                # Upload Resumes mode
                st.markdown("**Upload resume files (optional):**")
                uploaded_files = st.file_uploader(
                    "Upload Resume Files",
                    type=["pdf", "docx", "txt"],
                    accept_multiple_files=True,
                    help="Optional: Upload candidate resumes in PDF, DOCX, or TXT format. These will be automatically saved to your database after analysis.",
                    key="resume_uploader"
                )

                if uploaded_files:
                    st.success(f"Uploaded {len(uploaded_files)} resume(s)")
                    with st.expander("View uploaded files", expanded=False):
                        for file in uploaded_files:
                            size_kb = file.size / 1024
                            st.write(f"- {file.name} ({size_kb:.1f} KB)")
            
            # Show combined total
            total_candidates = len(selected_candidate_ids) + (len(uploaded_files) if uploaded_files else 0)
            if total_candidates > 0:
                source_breakdown = []
                if selected_candidate_ids:
                    source_breakdown.append(f"{len(selected_candidate_ids)} from database")
                if uploaded_files:
                    source_breakdown.append(f"{len(uploaded_files)} uploaded")
                breakdown_text = ", ".join(source_breakdown) if source_breakdown else "0"
                st.info(f"**Total candidates for analysis:** {total_candidates} ({breakdown_text})")
            else:
                st.info("**Info:** No resumes selected. The analysis will be created with job requirements only. You can add candidates later or upload resumes now.")

            # Process Button
            st.markdown('<div class="section-header"></div>', unsafe_allow_html=True)

            col1, col2, col3 = st.columns([1, 1, 1])

            with col2:
                # Use a custom styled button with brand green
                process_button = st.button(
                    "Process Candidates",
                    width='stretch',
                    type="primary",
                    key="process_btn"
                )

            # Process - either button was clicked OR processing was started and needs to continue
            # Check if processing was started on a previous rerun (button was clicked, rerun triggered, now continue)
            should_process = process_button or (st.session_state.get('processing', False) and st.session_state.get('processing_started', False))
            
            if should_process:
                # Only validate if button was just clicked (not on rerun continuation)
                if process_button:
                    # Validation
                    errors = []
                    
                    if not job_title:
                        errors.append("Job title is required")
                    if not location:
                        errors.append("Location is required")
                    if not job_description:
                        errors.append("Job description is required")
                    # Note: Resumes are now optional - removed validation for them

                    if errors:
                        for error in errors:
                            st.error(error)
                        # Clear processing flags if validation fails
                        st.session_state['processing'] = False
                        st.session_state['processing_started'] = False
                    else:
                        # Validation passed - set processing flag
                        st.session_state['processing'] = True
                        st.session_state['show_results_page'] = False
                        
                        # Initialize progress tracking
                        st.session_state['loading_progress'] = {
                            'step': 'analyzing',
                            'progress': 0.0,
                            'current': 0,
                            'total': 0
                        }
                
                # Continue with processing if processing flag is set (either from button click or rerun)
                if st.session_state.get('processing', False):
                    # Show loading screen immediately by setting processing flag and forcing a rerun
                    # This ensures the loading screen displays before long-running processing starts
                    # On first click: trigger rerun to show loading screen
                    # On rerun: processing_started will be True, so we continue with actual processing
                    if not st.session_state.get('processing_started', False):
                        st.session_state['processing_started'] = True
                        # Force immediate rerun to show loading screen
                        # On rerun, processing_started will be True, so we'll continue with processing
                        st.rerun()
                        st.stop()  # Stop execution to allow loading screen to render on rerun
                    
                    # Progress callback function for real-time updates
                    # This updates session state, and the loading screen in tab_new will read it
                    def update_progress(step, progress, current, total):
                        # Store progress in session state
                        st.session_state['loading_progress'] = {
                            'step': step,
                            'progress': progress,
                            'current': current,
                            'total': total
                        }
                    
                    # Process the candidates with progress tracking
                    # Note: The loading screen is shown in the tab_new section (line 2855)
                    # which reads from st.session_state['loading_progress']
                    # Progress updates are stored in session state and will be displayed
                    # when the loading screen renders (which happens immediately when processing starts)
                    start_time = datetime.now()

                    try:
                        # Save uploaded files temporarily and persist copies
                        temp_dir = tempfile.mkdtemp()
                        resume_paths = []
                        resume_assets = []
                        new_resume_data = []  # Store new resumes for database saving

                        try:
                            # Handle uploaded files
                            for uploaded_file in uploaded_files:
                                file_bytes = uploaded_file.getvalue()
                                file_path = os.path.join(temp_dir, uploaded_file.name)
                                with open(file_path, "wb") as f:
                                    f.write(file_bytes)
                                resume_paths.append(file_path)

                                stored_path, file_hash = save_bytes(file_bytes, uploaded_file.name)
                                resume_assets.append({
                                    "original_name": uploaded_file.name,
                                    "stored_path": stored_path,
                                    "file_hash": file_hash,
                                    "file_asset_id": None  # Will be set when file is saved to file_assets
                                })
                                
                                # Parse resume for potential database saving
                                from resume_parser import ResumeParser
                                parser = ResumeParser()
                                parsed_data = parser.parse(file_path)
                                new_resume_data.append({
                                    "parsed_data": parsed_data,
                                    "file_bytes": file_bytes,
                                    "original_name": uploaded_file.name,
                                    "stored_path": stored_path,
                                    "file_hash": file_hash
                                })
                            
                            # Handle selected candidates from database
                            if selected_candidate_ids:
                                selected_profiles = get_candidates_for_analysis(user_id, selected_candidate_ids)
                                for profile in selected_profiles:
                                    # Get resume file if available
                                    if profile.get('resume_file_id'):
                                        # Get stored path from file_assets
                                        with get_db() as conn:
                                            cur = conn.cursor()
                                            query = prepare_query(conn, """
                                                SELECT stored_path FROM file_assets
                                                WHERE id = ? AND user_id = ?
                                            """)
                                            cur.execute(query, (profile['resume_file_id'], user_id))
                                            row = cur.fetchone()
                                            if row and row[0]:
                                                stored_path = row[0]
                                                if os.path.exists(stored_path):
                                                    # Copy to temp directory for processing
                                                    import shutil
                                                    temp_file_path = os.path.join(temp_dir, f"{profile['name'] or 'candidate'}.pdf")
                                                    shutil.copy2(stored_path, temp_file_path)
                                                    resume_paths.append(temp_file_path)
                            
                            # Initialize app (logo is now fixed, no need to pass it)
                            app = CandidateRankerApp()

                            # Get edited skills if available (from file upload mode)
                            required_skills = st.session_state.get('current_required_skills', None)
                            preferred_skills = st.session_state.get('current_preferred_skills', None)
                            
                            # Get universal recruiting tool settings
                            industry_template = st.session_state.get('selected_template') or st.session_state.get('industry_template_select', None)
                            custom_weights = st.session_state.get('custom_weights', None)
                            dealbreakers = st.session_state.get('dealbreakers', [])
                            if 'dealbreakers_text' in st.session_state and st.session_state.get('dealbreakers_text'):
                                dealbreaker_text = st.session_state.get('dealbreakers_text', '')
                                dealbreakers = [line.strip() for line in dealbreaker_text.strip().split('\n') if line.strip()]
                            bias_reduction = st.session_state.get('bias_reduction_checkbox', False)
                            
                            # Run processing with progress callback
                            pdf_path = app.run(
                                job_title=job_title,
                                certifications=certifications,
                                location=location,
                                job_description=job_description,
                                resume_files=resume_paths,
                                required_skills=required_skills,
                                preferred_skills=preferred_skills,
                                progress_callback=update_progress,
                                user_id=st.session_state.get("user_id"),
                                resume_assets=resume_assets,
                                industry_template=industry_template,
                                custom_scoring_weights=custom_weights,
                                dealbreakers=dealbreakers if dealbreakers else None,
                                bias_reduction_enabled=bias_reduction
                            )
                            
                            # Get report ID if available (from app.run)
                            report_id = getattr(app, 'last_report_id', None)
                            
                            # Link selected database candidates to report
                            if report_id and selected_candidate_ids:
                                for candidate_id in selected_candidate_ids:
                                    link_candidate_to_analysis(candidate_id, report_id)
                            
                            # Automatically save new resumes to database
                            saved_count = 0
                            if new_resume_data and report_id:
                                import uuid
                                from db import get_db as db_get_db  # Import with alias to avoid scoping issues
                                
                                for resume_info in new_resume_data:
                                    try:
                                        # Save to file_assets
                                        asset_id = str(uuid.uuid4())
                                        with db_get_db() as conn:
                                            cur = conn.cursor()
                                            query = prepare_query(conn, """
                                                INSERT INTO file_assets 
                                                (id, user_id, kind, original_name, stored_path, created_at)
                                                VALUES (?, ?, ?, ?, ?, ?)
                                            """)
                                            cur.execute(query, (
                                                asset_id,
                                                user_id,
                                                'resume',
                                                resume_info['original_name'],
                                                resume_info['stored_path'],
                                                utcnow_str()
                                            ))
                                            try:
                                                conn.commit()
                                            except Exception as e:
                                                conn.rollback()
                                                logger.error(f"Failed to save file asset: {e}", exc_info=True)
                                                raise
                                        
                                        # Update resume_assets with file_asset_id for future reference
                                        for asset in resume_assets:
                                            if asset.get('original_name') == resume_info['original_name']:
                                                asset['file_asset_id'] = asset_id
                                                break
                                        
                                        # Save to candidate profiles
                                        profile_id = save_candidate_profile(
                                            user_id=user_id,
                                            resume_data=resume_info['parsed_data'],
                                            resume_file_id=asset_id,
                                            tags=[],
                                            notes=""
                                        )
                                        
                                        # Link to report
                                        link_candidate_to_analysis(profile_id, report_id)
                                        saved_count += 1
                                    except Exception as e:
                                        logger.error(f"Error auto-saving resume {resume_info['original_name']}: {e}", exc_info=True)
                                        # Continue with other resumes even if one fails
                                
                                # Store success message in session state for display
                                if saved_count > 0:
                                    st.session_state['resumes_auto_saved'] = saved_count

                            # Read PDF data before cleanup
                            with open(pdf_path, "rb") as f:
                                pdf_data = f.read()

                        finally:
                            # Cleanup temporary files
                            import shutil
                            try:
                                shutil.rmtree(temp_dir)
                            except Exception as e:
                                logger.warning(f"Failed to cleanup temp directory {temp_dir}: {e}", exc_info=True)
                                # Continue - cleanup errors are non-critical

                        # Calculate processing time
                        end_time = datetime.now()
                        processing_time = (end_time - start_time).total_seconds()

                        # Store results in session state
                        st.session_state.results = {
                            'candidate_scores': app.candidate_scores,
                            'pdf_path': pdf_path,
                            'pdf_data': pdf_data,
                            'job_details': {
                                'title': job_title,
                                'location': location,
                                'certifications': certifications
                            },
                            'job_details_obj': app.job_details,  # Store JobDetails object for dropdown
                            'processing_time': processing_time,
                            'timestamp': datetime.now()
                        }

                        # Clear processing flag and set results page flag
                        st.session_state['processing'] = False
                        st.session_state['show_results_page'] = True
                        # Clear loading progress
                        st.session_state.pop('loading_progress', None)

                        # Rerun to show results page
                        st.rerun()

                    except Exception as e:
                        # Clear processing flag on error
                        st.session_state['processing'] = False
                        st.session_state['show_results_page'] = False
                        st.session_state.pop('loading_progress', None)
                        st.session_state.pop('processing_started', None)
                        logger.error(f"Error processing candidates: {e}", exc_info=True)
                        st.error(f"An error occurred: {str(e)}")
                        st.exception(e)

            # Footer
            st.markdown("---")
            st.markdown(
                '<div style="text-align: center; color: #7f8c8d; padding: 1rem;">ResponsAble | Safety Staffing on Demand</div>',
                unsafe_allow_html=True
            )
        
        with tab_history:
            # Check if user wants to return to history list
            if st.session_state.get("show_history_list", False):
                st.session_state.pop("show_history_list", None)
                st.session_state.pop("viewing_previous_analysis", None)
                display_analysis_history(user_id)
            # Check if user wants to view a previous analysis
            elif st.session_state.get("view_report_id"):
                try:
                    analysis_data = load_analysis_data(st.session_state["view_report_id"], user_id)
                    if analysis_data:
                        # Store in results format for display_results
                        st.session_state.results = analysis_data
                        st.session_state["viewing_previous_analysis"] = True
                        st.session_state.pop("view_report_id", None)
                        st.rerun()
                    else:
                        st.error("**Warning:** Analysis not found or access denied. The report may have been deleted or you may not have permission to view it.")
                        logger.warning(f"Failed to load analysis {st.session_state['view_report_id']} for user {user_id}")
                        st.session_state.pop("view_report_id", None)
                        # Show the history list after error
                        display_analysis_history(user_id)
                except Exception as e:
                    logger.error(f"Error loading analysis data: {e}", exc_info=True)
                    st.error(f"**Error:** Error loading analysis: {str(e)}")
                    st.session_state.pop("view_report_id", None)
                    # Show the history list after error
                    display_analysis_history(user_id)
            # If we have stored results from viewing a previous analysis, show them
            elif st.session_state.results is not None and st.session_state.get("viewing_previous_analysis", False):
                display_results(st.session_state.results)
            else:
                display_analysis_history(user_id)
        
        with tab_database:
            display_resume_database(user_id)
        
        with tab_analytics:
            display_analysis_analytics(user_id)


if __name__ == "__main__":
    main()
